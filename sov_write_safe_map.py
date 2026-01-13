#!/usr/bin/env python3
import argparse, os, json, time, hashlib
from collections import defaultdict
from typing import List, Tuple, Optional, Dict

import openpyxl
from openpyxl.utils import get_column_letter

def sha256_file(path: str, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()

def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")

def find_code_headers(ws, max_row=150, max_col=80) -> List[Tuple[int,int]]:
    headers = []
    mr = min(ws.max_row or 0, max_row)
    mc = min(ws.max_column or 0, max_col)
    for r in range(1, mr+1):
        for c in range(1, mc+1):
            v = ws.cell(r,c).value
            if v is None:
                continue
            if str(v).strip().lower() == "code":
                headers.append((r,c))
    return headers

def is_digit_code(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        s = v.strip()
        # ignore formulas
        if s.startswith("="):
            return False
        return s.isdigit()
    # numeric types
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return v.is_integer()
    return False

def is_merged_cell(ws, row:int, col:int) -> bool:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return True
    return False

def main():
    ap = argparse.ArgumentParser(description="READ-ONLY: build write-safe column map for ESTIMATE (INPUT)")
    ap.add_argument("--template", required=True)
    ap.add_argument("--sheet", default="ESTIMATE (INPUT)")
    ap.add_argument("--scan-cols", default="A:AZ", help="Excel column range to scan, e.g. A:AZ")
    ap.add_argument("--scan-max-rows", type=int, default=2500)
    ap.add_argument("--out-dir", required=True)
    args = ap.parse_args()

    template = os.path.abspath(os.path.expanduser(args.template))
    out_dir = os.path.abspath(os.path.expanduser(args.out_dir))
    os.makedirs(out_dir, exist_ok=True)

    if not os.path.exists(template):
        raise SystemExit(f"FATAL: template not found: {template}")

    wb = openpyxl.load_workbook(template, data_only=False, keep_vba=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"FATAL: sheet not found: {args.sheet} (have: {wb.sheetnames})")
    ws = wb[args.sheet]

    headers = find_code_headers(ws)
    if not headers:
        raise SystemExit("FATAL: did not find a 'CODE' header in the governing sheet.")

    # Pick the leftmost CODE header (typical)
    headers_sorted = sorted(headers, key=lambda x: (x[1], x[0]))
    hr, hc = headers_sorted[0]

    # parse col range
    start_col_letter, end_col_letter = args.scan_cols.split(":")
    start_col_letter = start_col_letter.strip().upper()
    end_col_letter = end_col_letter.strip().upper()

    # convert to indices
    from openpyxl.utils.cell import column_index_from_string
    c1 = column_index_from_string(start_col_letter)
    c2 = column_index_from_string(end_col_letter)

    # find code rows under CODE header
    code_rows: List[int] = []
    blank_run = 0
    noncode_run = 0

    for r in range(hr+1, min(ws.max_row or 0, hr + args.scan_max_rows) + 1):
        v = ws.cell(r, hc).value
        if v is None or (isinstance(v, str) and not v.strip()):
            blank_run += 1
            if blank_run >= 25:
                break
            continue
        blank_run = 0

        if is_digit_code(v):
            code_rows.append(r)
            noncode_run = 0
        else:
            noncode_run += 1
            if noncode_run >= 5:
                break

    if not code_rows:
        raise SystemExit("FATAL: found CODE header but no numeric code rows below it.")

    # scan cells across code rows
    col_stats: Dict[int, Dict[str,int]] = defaultdict(lambda: {"formula":0, "merged":0, "blank":0, "value":0})
    for r in code_rows:
        for c in range(c1, c2+1):
            cell = ws.cell(r,c)
            if is_merged_cell(ws, r, c):
                col_stats[c]["merged"] += 1
                continue
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                col_stats[c]["formula"] += 1
                continue
            if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
                col_stats[c]["blank"] += 1
            else:
                col_stats[c]["value"] += 1

    safe_cols = []
    locked_cols = []
    for c in range(c1, c2+1):
        st = col_stats[c]
        if st["formula"] == 0 and st["merged"] == 0:
            safe_cols.append(get_column_letter(c))
        else:
            locked_cols.append(get_column_letter(c))

    report = {
        "timestamp": now_ts(),
        "template": template,
        "template_sha256": sha256_file(template),
        "governing_sheet": args.sheet,
        "code_header": {"row": hr, "col": hc, "col_letter": get_column_letter(hc)},
        "code_rows_count": len(code_rows),
        "code_rows_span": {"first": code_rows[0], "last": code_rows[-1]},
        "scan_cols": args.scan_cols,
        "safe_columns": safe_cols,
        "locked_columns": locked_cols,
        "column_stats": {
            get_column_letter(c): col_stats[c] for c in range(c1, c2+1)
        },
        "notes": [
            "SAFE = no formulas + no merged cells across all numeric code rows in ESTIMATE (INPUT).",
            "LOCKED = contains formula and/or merged cells in at least one code row.",
            "This is conservative by design."
        ]
    }

    base = os.path.splitext(os.path.basename(template))[0]
    safe_base = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in base)[:90]
    txt_path = os.path.join(out_dir, f"SOV_WriteSafeMap_{safe_base}.txt")
    json_path = os.path.join(out_dir, f"SOV_WriteSafeMap_{safe_base}.json")

    # TXT
    lines = []
    lines.append("SOV WRITE-SAFE MAP (READ-ONLY)")
    lines.append("")
    lines.append(f"Timestamp: {report['timestamp']}")
    lines.append(f"Template: {report['template']}")
    lines.append(f"SHA256: {report['template_sha256']}")
    lines.append(f"Governing sheet: {report['governing_sheet']}")
    lines.append(f"CODE header: R{hr}C{hc} ({get_column_letter(hc)})")
    lines.append(f"Code rows: {len(code_rows)} (R{code_rows[0]}..R{code_rows[-1]})")
    lines.append(f"Scan cols: {args.scan_cols}")
    lines.append("")
    lines.append(f"SAFE columns ({len(safe_cols)}): {', '.join(safe_cols)}")
    lines.append(f"LOCKED columns ({len(locked_cols)}): {', '.join(locked_cols)}")
    lines.append("")
    lines.append("Column stats (per code-row scan):")
    for col_letter in [get_column_letter(c) for c in range(c1, c2+1)]:
        st = report["column_stats"][col_letter]
        lines.append(f"- {col_letter}: formula={st['formula']} merged={st['merged']} blank={st['blank']} value={st['value']}")
    lines.append("")

    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)

    print("OK")
    print(f"TXT:  {txt_path}")
    print(f"JSON: {json_path}")
    print(f"SAFE columns: {len(safe_cols)}  LOCKED columns: {len(locked_cols)}")

if __name__ == "__main__":
    main()
