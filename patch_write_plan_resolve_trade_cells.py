#!/usr/bin/env python3
import json, re, sys, time
from pathlib import Path
from openpyxl import load_workbook

ADDR_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
CODE5_RE = re.compile(r"^\d{4,5}$")

def col_to_letters(n:int)->str:
    s=""
    while n>0:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def norm(s):
    return (s or "").strip().lower()

def find_header_row_and_col(ws, header_name, max_rows=200):
    target = norm(header_name)
    for r in range(1, max_rows+1):
        row_vals = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(r,c).value
            if v is None:
                continue
            sv = norm(str(v))
            if not sv:
                continue
            row_vals.append((c, sv))
        for c, sv in row_vals:
            if sv == target:
                return r, c
    return None, None

def cell_text(ws, r, c):
    v = ws.cell(r,c).value
    if v is None:
        return ""
    return str(v)

def row_blob(ws, r, max_cols=80):
    parts=[]
    for c in range(1, min(ws.max_column, max_cols)+1):
        v = ws.cell(r,c).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return " ".join(parts)

def is_code5(v):
    if v is None:
        return False
    s=str(v).strip()
    # handle floats like 1000.0 -> "1000"
    try:
        if isinstance(v, float) and v.is_integer():
            s = str(int(v))
    except Exception:
        pass
    return bool(CODE5_RE.match(s))

def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--plan", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--lock", required=True)
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--sheet", default="ESTIMATE (INPUT)")
    ap.add_argument("--code-header", default="CODE")
    ap.add_argument("--amount-header", default="AMOUNT")
    ap.add_argument("--scan-max-rows", type=int, default=2000)
    args = ap.parse_args()

    plan_path = Path(args.plan).expanduser()
    tpl_path  = Path(args.template).expanduser()
    lock_path = Path(args.lock).expanduser()
    out_dir   = Path(args.out_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)

    d = json.load(open(plan_path, "r", encoding="utf-8"))
    lk = json.load(open(lock_path, "r", encoding="utf-8"))

    sha = lk.get("template_sha256") or lk.get("template",{}).get("sha256")
    if not sha:
        raise SystemExit("FATAL: cannot find template_sha256 in lock JSON")
    d["template_sha256"] = sha

    wb = load_workbook(tpl_path, keep_vba=True, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"FATAL: sheet not found in template: {args.sheet}")
    ws = wb[args.sheet]

    hdr_row_code, col_code = find_header_row_and_col(ws, args.code_header)
    hdr_row_amt,  col_amt  = find_header_row_and_col(ws, args.amount_header)

    if col_amt is None:
        raise SystemExit(f"FATAL: could not find header '{args.amount_header}' in first 200 rows of {args.sheet}")
    if col_code is None:
        # not strictly required, but our matching uses it to only consider cost rows
        raise SystemExit(f"FATAL: could not find header '{args.code_header}' in first 200 rows of {args.sheet}")

    # choose a scan start row: below whichever header row is lower
    start_row = max(hdr_row_code or 1, hdr_row_amt or 1) + 1

    # build candidate rows that look like real cost-code rows (CODE is 5 digits)
    candidates=[]
    for r in range(start_row, min(start_row + args.scan_max_rows, ws.max_row)+1):
        if is_code5(ws.cell(r, col_code).value):
            blob = norm(row_blob(ws, r))
            code = str(ws.cell(r, col_code).value).strip()
            candidates.append((r, code, blob))

    if not candidates:
        raise SystemExit("FATAL: found zero 5-digit CODE rows to match against. Template structure unexpected.")

    writes = d.get("writes", [])
    if not isinstance(writes, list) or not writes:
        raise SystemExit("FATAL: plan has no writes[] to patch")

    problems=[]
    patched=0

    for i,w in enumerate(writes):
        if not isinstance(w, dict):
            problems.append((i, "write is not an object"))
            continue
        if norm(w.get("sheet")) != norm(args.sheet):
            problems.append((i, f"unexpected sheet '{w.get('sheet')}', expected '{args.sheet}'"))
            continue
        trade = None
        m = w.get("match")
        if isinstance(m, dict):
            trade = m.get("trade")
        trade_n = norm(trade)
        if not trade_n:
            problems.append((i, "missing match.trade"))
            continue

        # find rows whose blob contains the trade token
        hits = [(r, code) for (r, code, blob) in candidates if trade_n in blob]
        if len(hits) == 0:
            problems.append((i, f"no row match for trade='{trade}' (searched 4/5-digit CODE rows only)"))
            continue
        if len(hits) > 1:
            # too ambiguous: show top few
            preview = "; ".join([f"{code}@R{r}" for (r,code) in hits[:8]])
            problems.append((i, f"ambiguous trade='{trade}' matches={len(hits)} e.g. {preview}"))
            continue

        r, code = hits[0]
        cell = f"{col_to_letters(col_amt)}{r}"
        if not ADDR_RE.match(cell):
            problems.append((i, f"computed invalid cell {cell}"))
            continue

        # patch cell into the write object + nested write
        w["cell"] = cell
        if isinstance(w.get("write"), dict):
            w["write"]["cell"] = cell
        w.setdefault("meta", {})
        if isinstance(w["meta"], dict):
            w["meta"]["resolved_from"] = {"trade": trade, "code": code, "header_amount_col": col_amt, "row": r}
        patched += 1

    ts = time.strftime("%Y%m%d_%H%M%S")
    out = out_dir / f"{plan_path.stem}__CELLS_{ts}.json"
    json.dump(d, open(out, "w", encoding="utf-8"), indent=2)

    print("OK")
    print("IN_PLAN :", str(plan_path))
    print("OUT_PLAN:", str(out))
    print("PATCHED_WRITES:", patched, "/", len(writes))
    print("AMOUNT_COL:", col_to_letters(col_amt), f"(col {col_amt})", "CODE_COL:", col_to_letters(col_code), f"(col {col_code})")
    if problems:
        print("NO-GO")
        for idx, msg in problems:
            print(f"- writes[{idx}]: {msg}")
        sys.exit(2)
    print("GO")

if __name__ == "__main__":
    main()
