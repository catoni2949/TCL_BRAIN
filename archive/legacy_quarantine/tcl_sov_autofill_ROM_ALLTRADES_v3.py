#!/usr/bin/env python3
import argparse, json, datetime, re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Header targets (locked)
HDR_PROJECT   = "O2"
HDR_ADDR1     = "O4"
HDR_CITYSTZIP = "O5"
HDR_ESTIMATOR = "T3"
HDR_DATE      = "T2"

# SF cells (locked)
CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"
COST_SF_FORMULA = '=IFERROR(SUM($T:$T)/H4,0)'

# Conservative ROM $/SF by CODE (starter brain)
ROM_PER_SF_BY_CODE = {
    15300: 2.00,   # Fire Protection
    15400: 6.00,   # Plumbing
    15500: 8.00,   # HVAC
    16000: 9.00,   # Electrical base
    16500: 2.50,   # Lighting
    16700: 1.25,   # Data/Comms
    16950: 1.00,   # Fire Alarm
    9100:  10.00,  # GWB/Studs bucket
    7100:  3.00,   # Waterproofing bucket
    8100:  4.00,   # Doors bucket
    9900:  2.50,   # Paint bucket
    6400:  2.00,   # Casework placeholder
    11000: 3.00,   # Equipment placeholder
}
DEFAULT_PER_SF = 1.00  # any SUB row not listed above

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for rng in ws.merged_cells.ranges:
        if addr in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def _norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def find_header_row(ws, max_scan=120):
    max_c = min(ws.max_column, 120)
    for r in range(1, min(ws.max_row, max_scan) + 1):
        row = [_norm(ws.cell(r, c).value) for c in range(1, max_c + 1)]
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    raise RuntimeError("Could not find header row containing CODE + DESCRIPTION")

def find_col_exact(ws, header_row, name):
    name = _norm(name)
    for c in range(1, ws.max_column + 1):
        if _norm(ws.cell(header_row, c).value) == name:
            return c
    return None

def pick_real_desc_col(ws, header_row, desc_col):
    # Your template has DESCRIPTION header in col C but real text is in col B.
    # We detect this by sampling the next 60 rows and choosing the column with more non-empty strings.
    def score(col):
        s = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + 60) + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.strip():
                s += 1
        return s

    s_here = score(desc_col)
    s_left = score(desc_col - 1) if desc_col > 1 else -1
    s_right = score(desc_col + 1) if desc_col < ws.max_column else -1

    best_col = desc_col
    best = s_here
    if s_left > best:
        best, best_col = s_left, desc_col - 1
    if s_right > best:
        best, best_col = s_right, desc_col + 1

    return best_col

def pick_unitprice_subs_col(ws, header_row, unit_col):
    hours_col = find_col_exact(ws, header_row, "HOURS")
    subs_cols = [c for c in range(1, ws.max_column + 1) if _norm(ws.cell(header_row, c).value) == "SUBS"]
    cand = []
    for c in subs_cols:
        if c <= unit_col:
            continue
        if hours_col and c >= hours_col:
            continue
        cand.append(c)
    if cand:
        return cand[0]
    if subs_cols:
        return subs_cols[0]
    raise RuntimeError("Could not find any SUBS column in header row")

def load_overrides(path: Path) -> dict[int, float]:
    if not path or not path.exists():
        return {}
    out = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        try:
            code = int(float(k.strip()))
            amt = float(str(v).strip().replace(",", ""))
        except Exception:
            continue
        out[code] = amt
    return out

def write_headers(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value   = project
    _top_left_of_merge(ws, HDR_ADDR1).value     = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    ws[HDR_ESTIMATOR].value = "RNC"
    ws[HDR_DATE].value = datetime.date.today().strftime("%m/%d/%Y")

def write_sf(ws, total_sf):
    ws[CELL_TOTAL_SF].value = float(total_sf)
    ws[CELL_COST_SF].value = COST_SF_FORMULA

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--overrides", help="optional overrides file: code=amount per line")
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header = find_header_row(ws)

    code_col = find_col_exact(ws, header, "CODE")
    desc_col_hdr = find_col_exact(ws, header, "DESCRIPTION")
    unit_col = find_col_exact(ws, header, "UNIT")
    if not (code_col and desc_col_hdr and unit_col):
        raise RuntimeError(f"Header row {header} missing one of CODE/DESCRIPTION/UNIT. Found CODE={code_col} DESC={desc_col_hdr} UNIT={unit_col}")

    # FIX: choose the column that actually contains description strings
    desc_col = pick_real_desc_col(ws, header, desc_col_hdr)

    subs_col = pick_unitprice_subs_col(ws, header, unit_col)

    overrides = load_overrides(Path(args.overrides)) if args.overrides else {}

    write_headers(ws, args.project, args.addr1, args.citystzip)
    write_sf(ws, args.sf)

    wrote = []
    seen_codes = set()

    for r in range(header + 1, ws.max_row + 1):
        unit = _norm(ws.cell(r, unit_col).value)
        if unit != "SUB":
            continue

        raw_code = ws.cell(r, code_col).value
        try:
            code = int(float(raw_code))
        except Exception:
            continue

        if code in seen_codes:
            continue

        desc = ws.cell(r, desc_col).value
        if not (isinstance(desc, str) and desc.strip()):
            continue

        target = ws.cell(r, subs_col)
        if isinstance(target, MergedCell):
            continue
        if _is_formula(target):
            continue
        if _is_colored(target):
            continue

        if code in overrides:
            amt = overrides[code]
            src = "OVERRIDE"
        else:
            rate = ROM_PER_SF_BY_CODE.get(code, DEFAULT_PER_SF)
            amt = float(rate) * float(args.sf)
            src = f"ROM_${rate:.2f}/SF"

        target.value = float(amt)
        wrote.append({"code": code, "row": r, "subs_col": subs_col, "amount": amt, "src": src, "desc": desc.strip()})
        seen_codes.add(code)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(".rom_writes.json")
    sidecar.write_text(json.dumps({"writes": wrote}, indent=2), encoding="utf-8")

    print("SUCCESS")
    print("HEADER_ROW:", header, "CODE_COL:", code_col, "DESC_COL_HDR:", desc_col_hdr, "DESC_COL_USED:", desc_col, "UNIT_COL:", unit_col, "SUBS_COL:", subs_col)
    print("ROM WRITES:", len(wrote))
    print("OUTPUT:", str(out))
    print("SIDECAR:", str(sidecar))

if __name__ == "__main__":
    main()
