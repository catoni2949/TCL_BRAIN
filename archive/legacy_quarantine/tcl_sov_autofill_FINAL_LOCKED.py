#!/usr/bin/env python3
import argparse, json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Locked columns (confirmed from your scans)
COL_CODE = 1   # A
COL_DESC = 2   # B  (real descriptions live here)
COL_UNIT = 5   # E  (has "SUB" on the sub rows)
COL_SUBS = 9   # I  (first SUBS column in UNIT PRICES)

# Locked header cells (confirmed)
HDR_PROJECT   = "O3"
HDR_ADDR1     = "O4"
HDR_CITYSTZIP = "O5"
HDR_ESTIMATOR = "U3"   # RNC

CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"

def _norm(v):
    return str(v or "").strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def _guard_can_write(cell, label, *, allow_colored=False, allow_formula=False):
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"{label}: target cell is merged (read-only).")
    if (not allow_colored) and _is_colored(cell):
        raise RuntimeError(f"{label}: target cell is colored/protected.")
    if (not allow_formula) and _is_formula(cell):
        raise RuntimeError(f"{label}: target cell is a formula (won't overwrite).")

def write_header(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value = project
    _top_left_of_merge(ws, HDR_ADDR1).value = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    _top_left_of_merge(ws, HDR_ESTIMATOR).value = "RNC"
    # blank the next line so old values never linger
    try:
        _top_left_of_merge(ws, "O6").value = ""
    except Exception:
        pass
    # IMPORTANT: leave DATE alone (per your instruction)

def write_total_sf_and_rate(ws, total_sf):
    # H4 is a colored input cell in your template => allow colored overwrite
    h4 = ws[CELL_TOTAL_SF]
    _guard_can_write(h4, CELL_TOTAL_SF, allow_colored=True, allow_formula=False)
    h4.value = float(total_sf)

    # H6 we force formula every time; allow overwrite even if colored or currently a number/formula
    h6 = ws[CELL_COST_SF]
    _guard_can_write(h6, CELL_COST_SF, allow_colored=True, allow_formula=True)
    h6.value = "=IFERROR(SUM($T:$T)/H4,0)"

def load_amounts_file(path: Path):
    # amounts.json format:
    # 16000=105097
    # 15500=99782
    amounts = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            raise RuntimeError(f"Bad line in {path}: {line!r} (need CODE=AMOUNT)")
        k, v = line.split("=", 1)
        code = int(float(k.strip()))
        amt = float(v.strip().replace(",", ""))
        amounts[code] = amt
    return amounts

def find_target_row(ws, code_int):
    for r in range(1, ws.max_row + 1):
        try:
            c = ws.cell(r, COL_CODE).value
            if c is None:
                continue
            c_int = int(float(c))
        except Exception:
            continue
        if c_int != code_int:
            continue

        unit = _norm(ws.cell(r, COL_UNIT).value)
        if unit != "SUB":
            continue

        subs_cell = ws.cell(r, COL_SUBS)
        # for SUBS cells we keep guards strict (no merged/colored/formula)
        if isinstance(subs_cell, MergedCell) or _is_colored(subs_cell) or _is_formula(subs_cell):
            continue

        return r
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)

    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", default=None, help="Path to amounts.json (lines: CODE=AMOUNT)")
    ap.add_argument("--elec", type=float, default=None)
    ap.add_argument("--mech", type=float, default=None)

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    write_header(ws, args.project, args.addr1, args.citystzip)
    write_total_sf_and_rate(ws, args.sf)

    amounts = {}
    if args.amounts:
        amounts.update(load_amounts_file(Path(args.amounts)))
    if args.elec is not None:
        amounts[16000] = float(args.elec)
    if args.mech is not None:
        amounts[15500] = float(args.mech)

    wrote = []
    warnings = []

    for code_int, amt in sorted(amounts.items()):
        row = find_target_row(ws, code_int)
        if row is None:
            warnings.append({"code": code_int, "warn": "no_target_row_found"})
            continue

        cell = ws.cell(row, COL_SUBS)
        try:
            _guard_can_write(cell, f"SUBS row {row} (code {code_int})", allow_colored=False, allow_formula=False)
            cell.value = float(amt)
            wrote.append({"code": code_int, "row": row, "subs_col": COL_SUBS, "amount": float(amt), "desc": ws.cell(row, COL_DESC).value})
        except Exception as e:
            warnings.append({"code": code_int, "row": row, "warn": str(e)})

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(".warnings.json")
    sidecar.write_text(json.dumps({"wrote": wrote, "warnings": warnings}, indent=2), encoding="utf-8")

    print("SUCCESS")
    print("OUTPUT:", str(out))
    print("WROTE_COUNT:", len(wrote), "WARNINGS_COUNT:", len(warnings))
    if warnings:
        print("WARN_FILE:", str(sidecar))

if __name__ == "__main__":
    main()
