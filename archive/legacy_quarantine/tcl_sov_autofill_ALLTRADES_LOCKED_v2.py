#!/usr/bin/env python3
import argparse
import re
from pathlib import Path
from datetime import date

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# ---- LOCKED HEADER TARGETS (based on your scan of the template) ----
HDR_PROJECT   = "O2"   # PROJECT label is at N2; value cell is O2 (merged)
HDR_ADDR1     = "O4"   # LOCATION label at N4; line 1 goes O4
HDR_CITYSTZIP = "O5"   # line 2 goes O5
HDR_ESTIMATOR = "T3"   # ESTIMATOR label at S3; value cell is T3

# ---- SF / COST-SF cells (your confirmed locations) ----
CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"
COST_SF_FORMULA = '=IFERROR(SUM($T:$T)/H4,0)'

# ---- LOCKED TABLE COLS (based on your scans) ----
COL_CODE = 1   # A
COL_DESC = 2   # B  (real descriptions live here)
COL_SUBS = 9   # I  (first SUBS column in UNIT PRICES)

def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip().upper()

def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell) -> bool:
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for rng in ws.merged_cells.ranges:
        if addr in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def write_header(ws, project, addr1, citystzip):
    # IMPORTANT: DO NOT TOUCH DATE (T2) AT ALL.
    _top_left_of_merge(ws, HDR_PROJECT).value   = project
    _top_left_of_merge(ws, HDR_ADDR1).value     = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    _top_left_of_merge(ws, HDR_ESTIMATOR).value = "RNC"

def write_total_sf_and_cost_sf(ws, total_sf: float):
    # You told me H4 is an input, even if it’s colored. So we allow writing H4.
    h4 = ws[CELL_TOTAL_SF]
    h4.value = float(total_sf)

    # H6 must be formula (we overwrite whatever’s there)
    h6 = ws[CELL_COST_SF]
    if isinstance(h6, MergedCell):
        raise RuntimeError(f"{CELL_COST_SF} is merged; cannot write formula.")
    h6.value = COST_SF_FORMULA

def read_amounts_kv(path: Path) -> dict[int, float]:
    """
    Supports:
      16000=105097
      15500=99782
    Ignores blank lines and # comments.
    """
    m = {}
    txt = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    for line in txt:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            raise RuntimeError(f"Bad amounts line (missing '='): {line!r}")
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not k:
            continue
        try:
            code = int(float(k))
        except Exception:
            raise RuntimeError(f"Bad code in amounts: {k!r}")
        try:
            amt = float(str(v).replace(",", ""))
        except Exception:
            raise RuntimeError(f"Bad amount in amounts for {code}: {v!r}")
        m[code] = amt
    return m

def find_first_row_for_code(ws, code: int) -> int | None:
    """
    Picks the first “real” row for that code:
    - CODE matches in col A
    - DESCRIPTION in col B is not blank
    - Avoids obvious rollup/total rows by skipping descriptions that are exactly the trade name / very short blanky
    """
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, COL_CODE).value
        try:
            ci = int(float(v))
        except Exception:
            continue
        if ci != code:
            continue

        desc = ws.cell(r, COL_DESC).value
        d = _norm(desc)
        if not d:
            continue

        # Skip rows that look like rollups/totals (common pattern: same code + trade name only)
        if d in ("ELECTRICAL", "HVAC", "FIRE PROTECTION", "PLUMBING", "DRYWALL", "PAINTING", "METALS", "CONCRETE"):
            continue

        # Must have a writable SUBS cell in col I (non-merged)
        subs_cell = ws.cell(r, COL_SUBS)
        if isinstance(subs_cell, MergedCell):
            continue

        return r

    return None

def guarded_write_subs(ws, row: int, amount: float) -> None:
    cell = ws.cell(row=row, column=COL_SUBS)

    # Never touch formula cells; never touch colored “total” rows (but normal rows are fine)
    if _is_formula(cell):
        raise RuntimeError(f"Row {row} col {COL_SUBS}: is a formula; refusing.")
    if _is_colored(cell):
        raise RuntimeError(f"Row {row} col {COL_SUBS}: is colored/protected; refusing.")

    cell.value = float(amount)

def make_starter_amounts(ws, out_path: Path):
    """
    Writes a starter amounts file with all distinct CODEs that appear to be subcontract unit-price rows
    (based on having 'SUB' in UNIT column D on the first occurrence).
    It writes zeros so you can quickly replace numbers later.
    """
    # Find likely UNIT column: in your table it’s D (4)
    COL_UNIT = 4

    seen = set()
    codes = []

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, COL_CODE).value
        try:
            ci = int(float(v))
        except Exception:
            continue
        if ci in seen:
            continue

        unit = _norm(ws.cell(r, COL_UNIT).value)
        desc = _norm(ws.cell(r, COL_DESC).value)
        if unit == "SUB" and desc:
            seen.add(ci)
            codes.append(ci)

    codes.sort()
    lines = ["# code=amount  (edit amounts on the right side)\n"]
    for c in codes:
        lines.append(f"{c}=0\n")

    out_path.write_text("".join(lines), encoding="utf-8")
    print("WROTE STARTER AMOUNTS:", str(out_path))
    print("CODES:", len(codes))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", help="Path to amounts file in code=amount format")
    ap.add_argument("--make_amounts", action="store_true", help="Generate starter amounts file and exit")

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    # Option: generate starter amounts and exit
    if args.make_amounts:
        outp = Path(args.amounts or "amounts.json")
        make_starter_amounts(ws, outp)
        return

    if not args.amounts:
        raise RuntimeError("--amounts is required unless --make_amounts is used.")

    amounts = read_amounts_kv(Path(args.amounts))

    # Header + SF (DO NOT WRITE DATE)
    write_header(ws, args.project, args.addr1, args.citystzip)
    write_total_sf_and_cost_sf(ws, args.sf)

    wrote = []
    for code, amt in sorted(amounts.items()):
        if amt is None:
            continue
        try:
            if float(amt) == 0.0:
                continue
        except Exception:
            continue

        row = find_first_row_for_code(ws, int(code))
        if row is None:
            continue
        guarded_write_subs(ws, row, float(amt))
        wrote.append((int(code), row, float(amt), ws.cell(row, COL_DESC).value))

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print("SUCCESS")
    print("PROJECT ->", HDR_PROJECT, " | LOCATION ->", HDR_ADDR1, "/", HDR_CITYSTZIP, " | ESTIMATOR ->", HDR_ESTIMATOR, "(RNC)")
    print("TOTAL_SF", CELL_TOTAL_SF, "=", float(args.sf))
    print("COST/SF ", CELL_COST_SF, "=", COST_SF_FORMULA)
    print("ROM WRITES:", len(wrote))
    for code, row, amt, desc in wrote[:12]:
        print(f"  {code} row {row}  I{row}={amt}  desc={desc!r}")
    if len(wrote) > 12:
        print("  ...")
    print("OUTPUT:", str(out))

if __name__ == "__main__":
    main()
