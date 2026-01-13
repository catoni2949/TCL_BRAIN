#!/usr/bin/env python3
import argparse
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

SHEET = "ESTIMATE (INPUT)"

# Locked (confirmed)
CELL_PROJECT   = "O2"
CELL_ADDR1     = "O4"
CELL_CITYSTZIP = "O5"
CELL_DATE      = "T2"
CELL_ESTIMATOR = "T3"
CELL_TOTAL_SF  = "H4"
CELL_COST_SF   = "H6"

# Columns we know exist, but we will DETECT which ones are really CODE/DESC.
# Defaults are safe for your template.
DEFAULT_COL_CODE = 1   # A
DEFAULT_COL_DESC_USED = 2  # B (your real descriptions live here)
COL_UNIT = 5   # E
COL_LBR  = 7   # G
COL_MTL  = 8   # H
COL_SUBS = 9   # I  (first SUBS column)

YELLOW = PatternFill(patternType="solid", fgColor="FFF2CC")

# Self-perform: block SUBS by default -> split total into LBR+MTL
SELF_PERFORM_BLOCK_SUBS = {
    1010,1020,1030,1040,1050,1060,1080,
    1000,
    1600,1610,1630,1640,1680,1690,
    1710,1711,1715,1730,1750,1760,
    1700,
    2090,2100,2200,2900,
    2000,
    3000,3010,3300,3550,3710,
    5000,5100,5500,
    6100,6200,6300,
}

# Allow BOTH LBR + SUBS (install is self-performed, buyout still SUBS)
SPECIAL_ALLOW_SUBS_AND_LBR = {8750, 10500, 10800}

def _norm(v):
    return str(v or "").strip().upper()

def int_money(x):
    try:
        return int(round(float(x)))
    except Exception:
        return 0

def safe_cell(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    return ws[addr]

def safe_write(ws, addr, value):
    c = safe_cell(ws, addr)
    c.value = value

def set_date_preserve_excel(ws):
    c = safe_cell(ws, CELL_DATE)
    fmt = c.number_format
    c.value = date.today()
    if fmt:
        c.number_format = fmt

def write_total_sf_and_rate(ws, sf):
    ws[CELL_TOTAL_SF].value = int_money(sf)
    ws[CELL_COST_SF].value = "=IFERROR(SUM($T:$T)/H4,0)"

def find_header_row_and_cols(ws, max_scan_rows=220, max_scan_cols=60):
    """
    Finds header row by scanning the whole row for CODE + DESCRIPTION,
    then determines which columns are used for CODE and DESCRIPTION header,
    and also selects the *real* description column used in data rows (B vs C).
    """
    header_row = None
    for r in range(1, max_scan_rows + 1):
        row_vals = [_norm(ws.cell(r, c).value) for c in range(1, max_scan_cols + 1)]
        if "CODE" in row_vals and "DESCRIPTION" in row_vals:
            header_row = r
            break

    if not header_row:
        return None, None, None, None

    row_vals = [_norm(ws.cell(header_row, c).value) for c in range(1, max_scan_cols + 1)]
    code_col = row_vals.index("CODE") + 1
    desc_hdr_col = row_vals.index("DESCRIPTION") + 1

    # Decide which column actually contains description TEXT in the line items.
    # In your template, header says DESCRIPTION in col C but real text is in col B.
    cand_cols = [2, desc_hdr_col]  # prefer B, but test both
    best_col = cand_cols[0]
    best_hits = -1

    for col in cand_cols:
        hits = 0
        for rr in range(header_row + 1, min(ws.max_row, header_row + 40) + 1):
            v = ws.cell(rr, col).value
            if isinstance(v, str) and len(v.strip()) >= 3:
                hits += 1
        if hits > best_hits:
            best_hits = hits
            best_col = col

    desc_used_col = best_col
    return header_row, code_col, desc_hdr_col, desc_used_col

def iter_code_rows(ws, header_row, code_col, desc_used_col):
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, code_col).value
        try:
            code = int(float(v))
        except Exception:
            continue
        desc = ws.cell(r, desc_used_col).value
        unit = ws.cell(r, COL_UNIT).value
        yield r, code, desc, unit

def pick_write_row(ws, header_row, code_col, desc_used_col, code):
    for r, c, desc, unit in iter_code_rows(ws, header_row, code_col, desc_used_col):
        if c != code:
            continue
        if _norm(desc) and _norm(desc) != "DESCRIPTION":
            return r
    return None

def ensure_rom_log(wb):
    name = "ROM_LOG"
    if name in wb.sheetnames:
        sh = wb[name]
    else:
        sh = wb.create_sheet(name)
        sh.append(["CODE","DESC","ROW","COL","AMOUNT","BASIS","ASSUMPTIONS","CONFIDENCE","SOURCE","DATE"])
    sh.sheet_state = "hidden"
    return sh

def add_inline_note(cell, basis, conf, scope_line):
    txt = f"ROM\nBasis: {basis}\nScope: {scope_line}\nConf: {conf}"
    cell.comment = Comment(txt, "TCL-ROM")
    cell.comment.width = 250
    cell.comment.height = 90

def write_money_cell(ws, row, col, amount, *, is_rom, note=None):
    cell = ws.cell(row=row, column=col)
    cell.value = int_money(amount)
    if is_rom:
        cell.fill = YELLOW
        if note:
            add_inline_note(cell, note["basis"], note["conf"], note["scope_line"])
    return True

def estimate_rom_amounts(sf):
    """
    Estimator brain v1 (SF-driven conservative TI ROM).
    """
    sf = float(sf)
    out = {}
    out[15500] = sf * 8.0   # HVAC/Mech
    out[16000] = sf * 8.5   # Electrical
    out[9100]  = sf * 1.0
    out[9250]  = sf * 2.5
    out[9600]  = sf * 3.0
    out[9900]  = sf * 2.0
    out[15300] = sf * 1.5
    out[15400] = sf * 2.5
    return {k: int_money(v) for k, v in out.items()}

def parse_kv_file(path: Path):
    if not path.exists():
        raise RuntimeError(f"Overrides file not found: {path}")
    out = {}
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not k:
            continue
        code = int(float(k))
        out[code] = int_money(v)
    return out

def apply_amount(ws, header_row, code_col, desc_used_col, code, total, *,
                 is_rom, self_lbr_pct, rom_writes, override_writes, skipped_self_perf, rom_log):
    if int_money(total) == 0:
        return

    row = pick_write_row(ws, header_row, code_col, desc_used_col, code)
    if row is None:
        return

    desc = ws.cell(row=row, column=desc_used_col).value or ""
    scope_line = (str(desc)[:45] + "…") if len(str(desc)) > 45 else str(desc)

    conf = "MED" if code in (15500,16000) else "LOW"
    basis = "SF-driven TI ROM"
    source = "Inferred"

    note = {"basis": basis, "conf": conf, "scope_line": scope_line}

    def log(colname, amt):
        rom_log.append([code, str(desc), row, colname, int_money(amt), basis, "", conf, source, str(date.today())])

    allow_subs = True
    if (code in SELF_PERFORM_BLOCK_SUBS) and (code not in SPECIAL_ALLOW_SUBS_AND_LBR):
        allow_subs = False

    if not allow_subs:
        total_i = int_money(total)
        lbr_i = int(round(total_i * float(self_lbr_pct)))
        mtl_i = total_i - lbr_i
        write_money_cell(ws, row, COL_LBR, lbr_i, is_rom=is_rom, note=note if is_rom else None); log("LBR", lbr_i)
        write_money_cell(ws, row, COL_MTL, mtl_i, is_rom=is_rom, note=note if is_rom else None); log("MTL", mtl_i)
        (rom_writes if is_rom else override_writes).append({"code": code, "row": row})
    else:
        write_money_cell(ws, row, COL_SUBS, total, is_rom=is_rom, note=note if is_rom else None); log("SUBS", total)
        (rom_writes if is_rom else override_writes).append({"code": code, "row": row})

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--estimator", default="RNC")
    ap.add_argument("--overrides", default=None)
    ap.add_argument("--self_lbr_pct", type=float, default=0.60)
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'")
    ws = wb[SHEET]

    safe_write(ws, CELL_PROJECT, args.project)
    safe_write(ws, CELL_ADDR1, args.addr1)
    safe_write(ws, CELL_CITYSTZIP, args.citystzip)

    set_date_preserve_excel(ws)
    safe_write(ws, CELL_ESTIMATOR, args.estimator)

    write_total_sf_and_rate(ws, args.sf)

    header_row, code_col, desc_hdr_col, desc_used_col = find_header_row_and_cols(ws)
    if not header_row:
        raise RuntimeError("Could not find header row (CODE/DESCRIPTION)")

    rom_log = ensure_rom_log(wb)

    rom_amounts = estimate_rom_amounts(args.sf)
    overrides = parse_kv_file(Path(args.overrides)) if args.overrides else {}

    rom_writes = []
    override_writes = []
    skipped_self_perf = []

    for code in sorted(rom_amounts.keys()):
        apply_amount(ws, header_row, code_col, desc_used_col, code, rom_amounts[code],
                     is_rom=True, self_lbr_pct=args.self_lbr_pct,
                     rom_writes=rom_writes, override_writes=override_writes,
                     skipped_self_perf=skipped_self_perf, rom_log=rom_log)

    for code in sorted(overrides.keys()):
        apply_amount(ws, header_row, code_col, desc_used_col, code, overrides[code],
                     is_rom=False, self_lbr_pct=args.self_lbr_pct,
                     rom_writes=rom_writes, override_writes=override_writes,
                     skipped_self_perf=skipped_self_perf, rom_log=rom_log)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(
        __import__("json").dumps(
            {
                "rom_writes": rom_writes,
                "override_writes": override_writes,
                "skipped_self_perf_subs": skipped_self_perf,
            },
            indent=2
        ),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"HEADER_ROW: {header_row}  CODE_COL:{code_col}  DESC_HDR_COL:{desc_hdr_col}  DESC_USED_COL:{desc_used_col}")
    print(f"PROJECT cell: {CELL_PROJECT}")
    print(f"DATE cell: {CELL_DATE} = {ws[CELL_DATE].value!r}  format={ws[CELL_DATE].number_format!r}")
    print(f"ESTIMATOR cell: {CELL_ESTIMATOR} = {args.estimator}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int_money(args.sf)}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow + inline note + ROM_LOG)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"SKIPPED (self-perform SUBS block): {len(skipped_self_perf)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
