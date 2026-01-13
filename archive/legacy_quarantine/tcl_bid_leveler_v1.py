#!/usr/bin/env python3
"""
TCL Bid Leveling Engine v1 (B2 mode: completeness weighted over price)

Inputs:
  --bids_json  Parsed bids JSON (from tcl_bid_parser_v1.py)
  --truth_json Truth map JSON (optional; v1 uses it for light discipline flags only)
Outputs:
  --out_dir    Writes:
     bid_leveling.xlsx
     recommendations.json

Design notes:
- Warn-only engine: never blocks output; pushes issues into recommendations/clarifications.
- Completeness > price (B2): winner selection prioritizes scope completeness, then price.
"""
import argparse, json, os, re, datetime
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

DEFAULT_CANON_TOKENS = {"lachini", "vineyards", "tasting", "room"}

def safe_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*:/\\\?]', '-', name)
    return name[:31] if len(name) > 31 else name

def project_match_score(bid: dict, canon_tokens=set(DEFAULT_CANON_TOKENS)) -> float:
    text = (bid.get("project_name_raw") or "") + " " + (bid.get("project_address_raw") or "")
    toks = set(re.findall(r"[A-Za-z0-9]+", text.lower()))
    hits = len(toks & canon_tokens)
    if hits >= 2:
        return 1.0
    if hits == 1:
        return 0.6
    return 0.0

def completeness_score(bid: dict, canon_tokens=set(DEFAULT_CANON_TOKENS)) -> float:
    score = 70.0
    excl = len(bid.get("exclusions") or [])
    allow = len(bid.get("allowances") or [])
    alt = len(bid.get("alternates") or [])
    clar = len(bid.get("clarifications") or [])
    sched = len(bid.get("schedule_notes") or [])
    total = bid.get("base_total")

    if total is None:
        score -= 25

    score -= excl * 10
    score -= allow * 8

    # Alternates aren't inherently bad; only penalize if there are many
    if alt > 3:
        score -= (alt - 3) * 1.5

    if clar > 0:
        score += 5
    if sched > 0:
        score += 3

    score += project_match_score(bid, canon_tokens) * 10
    score += float(bid.get("confidence_score") or 0) * 10

    return max(0.0, min(100.0, score))

def trade_threshold(trade: str) -> float:
    # Tunable thresholds: some trades routinely carry allowances/exclusions but are still awardable.
    if trade.upper() == "FIRE ALARM":
        return 40.0
    if trade.upper() == "PLUMBING":
        return 50.0
    return 60.0

def qualifies(bid: dict, canon_tokens=set(DEFAULT_CANON_TOKENS)) -> bool:
    tr = (bid.get("trade") or "").upper()
    return (
        bid.get("base_total") is not None and
        completeness_score(bid, canon_tokens) >= trade_threshold(tr) and
        float(bid.get("confidence_score") or 0) >= 0.3
    )

def build_clarifications(bid: dict, canon_tokens=set(DEFAULT_CANON_TOKENS)) -> list[str]:
    qs = []
    bidder = bid.get("bidder_name") or "BIDDER"
    for ex in bid.get("exclusions") or []:
        qs.append(f"Confirm {bidder} exclusion: {ex}")
    for al in bid.get("allowances") or []:
        qs.append(f"Confirm {bidder} allowance covers: {al}")
    if bid.get("base_total") is None:
        qs.append(f"Provide base bid total for {bidder} ({bid.get('file')}).")
    if project_match_score(bid, canon_tokens) < 0.6:
        qs.append("Confirm bid is for the correct project (project name/address not clearly stated on proposal).")
    return qs

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bids_json", required=True, help="Parsed bids JSON produced by tcl_bid_parser_v1.py")
    ap.add_argument("--truth_json", default=None, help="Truth map JSON (optional; v1 uses light flags only)")
    ap.add_argument("--out_dir", required=True, help="Output directory")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    bids_doc = json.load(open(args.bids_json, "r", encoding="utf-8"))
    bids = bids_doc.get("bids", [])

    truth = None
    if args.truth_json and os.path.exists(args.truth_json):
        truth = json.load(open(args.truth_json, "r", encoding="utf-8"))

    # Canon tokens derived from bids (majority vote); fallback to defaults
    token_counts = {}
    for b in bids:
        n = b.get("project_name_raw") or ""
        for t in set(re.findall(r"[A-Za-z0-9]+", n.lower())):
            token_counts[t] = token_counts.get(t, 0) + 1
    top = sorted(token_counts.items(), key=lambda x: -x[1])[:10]
    canon_tokens = {t for t, c in top if c >= 2 and len(t) >= 4 and t not in {"name", "project", "job", "work"}}
    if not canon_tokens:
        canon_tokens = set(DEFAULT_CANON_TOKENS)

    # Group by trade
    groups = defaultdict(list)
    for b in bids:
        groups[(b.get("trade") or "UNKNOWN").upper()].append(b)

    # Recommendations
    summary_rows = []
    recommendations = []
    for trade, blist in groups.items():
        scored = [(b, completeness_score(b, canon_tokens), b.get("base_total")) for b in blist]
        qualified = [(b, cs, tot) for (b, cs, tot) in scored if qualifies(b, canon_tokens)]

        decision = {"trade": trade}
        if qualified:
            winner = sorted(qualified, key=lambda x: (-x[1], x[2]))[0]
            decision.update({
                "status": "RECOMMEND",
                "winner": winner[0].get("bidder_name"),
                "base_total": winner[2],
                "completeness": round(winner[1], 1),
            })
        else:
            present = [(b, cs, tot) for (b, cs, tot) in scored if tot is not None]
            if present:
                cw = sorted(present, key=lambda x: (-x[1], x[2]))[0]
                decision.update({
                    "status": "HOLD_CONDITIONAL",
                    "conditional_winner": cw[0].get("bidder_name"),
                    "base_total": cw[2],
                    "completeness": round(cw[1], 1),
                })
            else:
                decision.update({
                    "status": "HOLD_NEED_NUMBER"
                })

        summary_rows.append(decision)

        # Build recommendation package per trade
        trade_pkg = {
            "trade": trade,
            "decision": decision,
            "bids": []
        }
        for b in blist:
            trade_pkg["bids"].append({
                "bidder": b.get("bidder_name"),
                "file": b.get("file"),
                "project_raw": b.get("project_name_raw"),
                "total": b.get("base_total"),
                "confidence": b.get("confidence_score"),
                "project_match": project_match_score(b, canon_tokens),
                "completeness": round(completeness_score(b, canon_tokens), 1),
                "exclusions": b.get("exclusions") or [],
                "allowances": b.get("allowances") or [],
                "alternates": b.get("alternates") or [],
            })

        target = decision.get("winner") or decision.get("conditional_winner")
        if target:
            tbid = next((x for x in blist if (x.get("bidder_name") == target)), None)
            trade_pkg["clarifications"] = build_clarifications(tbid, canon_tokens) if tbid else []
        else:
            trade_pkg["clarifications"] = sorted({q for x in blist for q in build_clarifications(x, canon_tokens)})

        recommendations.append(trade_pkg)

    # Write Excel
    wb = Workbook()
    wb.remove(wb.active)

    summary_df = pd.DataFrame(summary_rows)
    ws = wb.create_sheet("SUMMARY")
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for trade, blist in groups.items():
        rows = []
        for b in blist:
            rows.append({
                "BIDDER": b.get("bidder_name"),
                "FILE": b.get("file"),
                "PROJECT_RAW": b.get("project_name_raw"),
                "TOTAL": b.get("base_total"),
                "CONF": b.get("confidence_score"),
                "PROJ_MATCH": project_match_score(b, canon_tokens),
                "COMPLETENESS": round(completeness_score(b, canon_tokens), 1),
                "#EXCL": len(b.get("exclusions") or []),
                "#ALLOW": len(b.get("allowances") or []),
                "#ALT": len(b.get("alternates") or []),
                "EXCLUSIONS": " | ".join((b.get("exclusions") or [])[:4]),
                "ALLOWANCES": " | ".join((b.get("allowances") or [])[:4]),
                "ALTERNATES": " | ".join((b.get("alternates") or [])[:4]),
            })
        df = pd.DataFrame(rows).sort_values(["COMPLETENESS", "TOTAL"], ascending=[False, True])
        ws = wb.create_sheet(safe_sheet_name(trade))
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_xlsx = os.path.join(args.out_dir, "bid_leveling.xlsx")
    wb.save(out_xlsx)

    out_json = os.path.join(args.out_dir, "recommendations.json")
    payload = {
        "generated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "canon_project_tokens": sorted(canon_tokens),
        "truth_map": truth,
        "recommendations": recommendations,
    }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print("OK")
    print("Wrote:", out_xlsx)
    print("Wrote:", out_json)

if __name__ == "__main__":
    main()
