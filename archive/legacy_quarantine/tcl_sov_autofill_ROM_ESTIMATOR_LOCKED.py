#!/usr/bin/env python3
import argparse
from datetime import date
from pathlib import Path
import json
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Header cells (locked/confirmed)
CELL_PROJECT   = "O2"
CELL_ADDR1     = "O4"
CELL_CITYSTZIP = "O5"
CELL_DATE      = "T2"
CELL_ESTIMATOR = "T3"

CELL_TOTAL_SF  = "H4"
CELL_COST_SF   = "H6"

# Column positions (locked/confirmed)
COL_CODE = 1    # A
COL_DESC = 2    # B (actual descriptions)
COL_UNIT = 5    # E
COL_LBR  = 7    # G
COL_MTL  = 8    # H
COL_SUBS = 9    # I

YELLOW_FILL = PatternFill(patternType="solid", fgColor="FFF2CC")  # light yellow

def _norm(v):
    return str(v or "").strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        # any solid/pattern fill = "colored"
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    return bool(rgb and rgb not in ("00000000", "FFFFFFFF"))

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def safe_write(ws, addr, value):
    c = _top_left_of_merge(ws, addr)
    c.value = value

def set_date_preserve_excel(ws):
    # write as real Excel date (not a string), preserve existing number format
    c = _top_left_of_merge(ws, CELL_DATE)
    fmt = c.number_format
    c.value = date.today()
    if fmt:
        c.number_format = fmt

def int_money(x):
    return int(round(float(x)))

def write_total_sf_and_rate(ws, sf):
    # H4 = total SF (number), H6 = cost/sf formula
    h4 = ws[CELL_TOTAL_SF]
    if isinstance(h4, MergedCell) or _is_formula(h4) or _is_colored(h4):
        # if template marks it colored, we still want to write; so only block if formula/merged
        if isinstance(h4, MergedCell) or _is_formula(h4):
            raise RuntimeError(f"{CELL_TOTAL_SF} is not writable (merged/formula).")
    h4.value = int_money(sf)

    # Cost/SF formula
    h6 = ws[CELL_COST_SF]
    if isinstance(h6, MergedCell):
        raise RuntimeError(f"{CELL_COST_SF} is merged; not writable.")
    # always set formula (template sometimes has a value)
    ws[CELL_COST_SF].value = f"=IFERROR(SUM($T:$T)/{CELL_TOTAL_SF},0)"

def find_header_row(ws):
    # find row where A has "CODE" and B has "DESCRIPTION" (or C in some templates)
    for r in range(1, 60):
        a = _norm(ws.cell(r, COL_CODE).value)
        b = _norm(ws.cell(r, COL_DESC).value)
        c = _norm(ws.cell(r, 3).value)  # sometimes DESCRIPTION label appears in C
        if a == "CODE" and (b == "DESCRIPTION" or c == "DESCRIPTION"):
            return r
    return None

def iter_code_rows(ws, header_row):
    # yield (row, code_int)
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(r, COL_CODE).value
        if v is None or str(v).strip() == "":
            continue
        try:
            code = int(float(v))
        except Exception:
            continue
        yield r, code

def cell_writable(cell):
    if isinstance(cell, MergedCell):
        return False
    if _is_formula(cell):
        return False
    if _is_colored(cell):
        return False
    return True

def find_write_row(ws, header_row, code):
    """
    Choose the first *non-colored, non-formula* row for this code, preferring
    rows that actually look like line items (has UNIT).
    """
    candidates = []
    for r, c in iter_code_rows(ws, header_row):
        if c != code:
            continue
        unit = _norm(ws.cell(r, COL_UNIT).value)
        # prefer rows with a unit (SUB/LS/SF/HR/etc.)
        score = 0
        if unit:
            score += 10
        # prefer if SUBS cell is writable (for subs-heavy trades)
        if cell_writable(ws.cell(r, COL_SUBS)):
            score += 2
        # avoid band/header rows by checking DESC exists
        if _norm(ws.cell(r, COL_DESC).value):
            score += 1
        candidates.append((score, r))

    if not candidates:
        return None

    candidates.sort(reverse=True)
    # now pick the first row where at least one of the target cells is writable
    for _, r in candidates:
        if (cell_writable(ws.cell(r, COL_SUBS)) or
            cell_writable(ws.cell(r, COL_LBR)) or
            cell_writable(ws.cell(r, COL_MTL))):
            return r
    return None

def add_rom_mark(ws, row, note_text):
    # yellow highlight on the written cells + comment on DESC
    desc_cell = ws.cell(row=row, column=COL_DESC)
    # keep sheet clean: short comment only
    if desc_cell.comment is None:
        desc_cell.comment = Comment(note_text, "TCL-ROM")
    else:
        # don't spam; append once
        if "TCL-ROM" not in (desc_cell.comment.author or ""):
            desc_cell.comment = Comment(note_text, "TCL-ROM")

def write_money_cell(ws, row, col, value, is_rom):
    cell = ws.cell(row=row, column=col)
    if not cell_writable(cell):
        return False
    cell.value = int_money(value)
    if is_rom:
        cell.fill = YELLOW_FILL
    return True

# ---- Your self-perform policy ----
SELF_PERFORM_BLOCK_SUBS = {
    # 01xxx
    1010,1020,1030,1040,1050,1060,1080,
    1000,
    # 016xx
    1600,1610,1630,1640,1680,1690,
    # 017xx
    1710,1711,1715,1730,1750,1760,
    1700,
    # 02xxx
    2090,2100,2200,2900,
    2000,
    # 03xxx
    3000,3010,3300,3550,3710,
    # 05xxx
    5000,5100,5500,
    # 06xxx (06400 stays buyout/subs allowed)
    6100,6200,6300,
}

# Special: allow SUBS buyout AND self-performed install (LBR/MTL)
SPECIAL_ALLOW_SUBS_AND_LBR = {8750,10500,10800}

def code5_to_int(code_str):
    # handle 01010 etc -> 1010
    s = re.sub(r"[^\d]", "", str(code_str))
    if not s:
        raise ValueError("bad code")
    return int(s)

# ---- ROM model (simple but useful TI defaults) ----
# Rates are $/SF. Keep conservative.
ROM_RATES_PER_SF = {
    # Interiors TI-ish placeholders (adjust later)
    9100: 10.0,   # steel studs & GWB
    9250: 9.0,    # drywall (if present as 09250 it becomes 9250)
    9500: 6.0,    # ACT
    9680: 10.0,   # flooring
    9900: 5.0,    # paint
    15400: 18.0,  # plumbing (TI rough placeholder)
    15300: 10.0,  # fire protection placeholder
    16000: 20.0,  # electrical base placeholder (often overridden)
    15500: 18.0,  # HVAC placeholder (often overridden)
}

# Fixed allowances ($) regardless of SF (used sparingly)
ROM_ALLOWANCES = {
    1630: 2500,   # permits & plans placeholder
}

def build_rom_amounts(total_sf):
    amounts = {}
    for code, rate in ROM_RATES_PER_SF.items():
        amounts[code] = int_money(rate * float(total_sf))
    for code, amt in ROM_ALLOWANCES.items():
        amounts[code] = int_money(amt)
    return amounts

def parse_kv_lines(lines):
    """
    Supports:
      15500=99782
      08750.subs=50000
      08750.lbr=12000
      08750.mtl=3000
    Returns dict: key -> value, where key is either int code or tuple(code, part)
    """
    out = {}
    for raw in lines:
        s = raw.strip()
        if not s or s.startswith("#"):
            continue
        if "=" not in s:
            continue
        k, v = s.split("=", 1)
        k = k.strip()
        v = v.strip().replace(",", "")
        if not v:
            continue
        try:
            amt = float(v)
        except Exception:
            continue

        part = None
        if "." in k:
            k0, part = k.split(".", 1)
            k0 = k0.strip()
            part = part.strip().lower()
            code = code5_to_int(k0)
            out[(code, part)] = amt
        else:
            code = code5_to_int(k)
            out[code] = amt
    return out

def parse_kv_file(path: Path):
    if not path.exists():
        raise RuntimeError(f"Overrides file not found: {path}")
    return parse_kv_lines(path.read_text(encoding="utf-8", errors="ignore").splitlines())

def apply_amount(ws, header_row, code, total,
                 self_lbr_pct,
                 is_rom,
                 rom_writes, override_writes, skipped_self_perf):
    if float(total) == 0:
        return

    row = find_write_row(ws, header_row, code)
    if row is None:
        return

    allow_subs = True
    wrote_any = False

    # If caller provided explicit part keys, those are handled elsewhere.
    # Here: "total" goes either to (LBR+MTL) for self-perform OR SUBS otherwise.
    if (code in SELF_PERFORM_BLOCK_SUBS) and (code not in SPECIAL_ALLOW_SUBS_AND_LBR):
        total_i = int_money(total)
        lbr_i = int(round(total_i * float(self_lbr_pct)))
        mtl_i = total_i - lbr_i
        ok1 = write_money_cell(ws, row, COL_LBR, lbr_i, is_rom=is_rom)
        ok2 = write_money_cell(ws, row, COL_MTL, mtl_i, is_rom=is_rom)
        wrote_any = ok1 or ok2
        if wrote_any and is_rom:
            add_rom_mark(ws, row, f"ROM: self-perform split {int(self_lbr_pct*100)}% LBR / {int((1-self_lbr_pct)*100)}% MTL")
    else:
        if allow_subs:
            wrote_any = write_money_cell(ws, row, COL_SUBS, total, is_rom=is_rom)
            if wrote_any and is_rom:
                add_rom_mark(ws, row, "ROM: no bid yet (SUBS placeholder)")
        else:
            skipped_self_perf.append(code)

    if wrote_any:
        (rom_writes if is_rom else override_writes).append({"code": code, "row": row, "total": int_money(total)})

def apply_explicit_parts(ws, header_row, code, parts_dict, is_rom, rom_writes, override_writes):
    row = find_write_row(ws, header_row, code)
    if row is None:
        return

    wrote_any = False
    if "lbr" in parts_dict:
        wrote_any |= write_money_cell(ws, row, COL_LBR, parts_dict["lbr"], is_rom=is_rom)
    if "mtl" in parts_dict:
        wrote_any |= write_money_cell(ws, row, COL_MTL, parts_dict["mtl"], is_rom=is_rom)
    if "subs" in parts_dict:
        wrote_any |= write_money_cell(ws, row, COL_SUBS, parts_dict["subs"], is_rom=is_rom)

    if wrote_any and is_rom:
        add_rom_mark(ws, row, "ROM: explicit parts (lbr/mtl/subs)")

    if wrote_any:
        (rom_writes if is_rom else override_writes).append({"code": code, "row": row, "parts": {k:int_money(v) for k,v in parts_dict.items()}})

def ensure_rom_log_sheet(wb):
    if "ROM_LOG" in wb.sheetnames:
        ws = wb["ROM_LOG"]
        return ws
    ws = wb.create_sheet("ROM_LOG")
    ws.append(["timestamp", "type", "code", "row", "desc", "lbr", "mtl", "subs", "basis"])
    return ws

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)

    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--estimator", default="RNC")

    # No amounts file required anymore
    ap.add_argument("--overrides", default=None, help="Overrides file (code=amt or code.subs/lbr/mtl=amt)")
    ap.add_argument("--override", action="append", default=[], help='Inline override, repeatable. e.g. --override "15500=99782"')

    ap.add_argument("--self_lbr_pct", type=float, default=0.60)

    args = ap.parse_args()

    wb = load_workbook(Path(args.template), keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)
    if not header_row:
        raise RuntimeError("Could not find header row (CODE/DESCRIPTION)")

    # Headers
    safe_write(ws, CELL_PROJECT, args.project)
    safe_write(ws, CELL_ADDR1, args.addr1)
    safe_write(ws, CELL_CITYSTZIP, args.citystzip)
    set_date_preserve_excel(ws)
    safe_write(ws, CELL_ESTIMATOR, args.estimator)

    # SF + cost/sf
    write_total_sf_and_rate(ws, args.sf)

    # Build internal ROM
    rom_amounts = build_rom_amounts(args.sf)

    # Load overrides (file + inline)
    overrides = {}
    if args.overrides:
        overrides.update(parse_kv_file(Path(args.overrides)))
    if args.override:
        overrides.update(parse_kv_lines(args.override))

    rom_writes = []
    override_writes = []
    skipped_self_perf = []

    romlog = ensure_rom_log_sheet(wb)
    ts = date.today().isoformat()

    # --- ROM pass (totals only) ---
    for code in sorted(rom_amounts.keys()):
        apply_amount(
            ws, header_row, code, rom_amounts[code],
            self_lbr_pct=args.self_lbr_pct,
            is_rom=True,
            rom_writes=rom_writes,
            override_writes=override_writes,
            skipped_self_perf=skipped_self_perf
        )

    # --- Overrides pass (wins) ---
    # First: group explicit part overrides
    part_overrides = {}
    simple_overrides = {}

    for k, v in overrides.items():
        if isinstance(k, tuple):
            code, part = k
            part_overrides.setdefault(code, {})[part] = v
        else:
            simple_overrides[k] = v

    # Apply simple overrides (treated like totals)
    for code in sorted(simple_overrides.keys()):
        apply_amount(
            ws, header_row, code, simple_overrides[code],
            self_lbr_pct=args.self_lbr_pct,
            is_rom=False,
            rom_writes=rom_writes,
            override_writes=override_writes,
            skipped_self_perf=skipped_self_perf
        )

    # Apply explicit part overrides (lbr/mtl/subs)
    for code in sorted(part_overrides.keys()):
        apply_explicit_parts(
            ws, header_row, code, part_overrides[code],
            is_rom=False,
            rom_writes=rom_writes,
            override_writes=override_writes
        )

    # ROM_LOG entries (only for ROM + overrides that actually wrote)
    def _desc_for_row(r):
        return ws.cell(r, COL_DESC).value

    for w in rom_writes:
        r = w["row"]
        romlog.append([ts, "ROM", w["code"], r, _desc_for_row(r),
                       ws.cell(r, COL_LBR).value, ws.cell(r, COL_MTL).value, ws.cell(r, COL_SUBS).value,
                       "internal ROM"])
    for w in override_writes:
        r = w["row"]
        romlog.append([ts, "OVERRIDE", w["code"], r, _desc_for_row(r),
                       ws.cell(r, COL_LBR).value, ws.cell(r, COL_MTL).value, ws.cell(r, COL_SUBS).value,
                       "user override"])

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(
        json.dumps(
            {
                "rom_writes": rom_writes,
                "override_writes": override_writes,
                "skipped_self_perf_subs": skipped_self_perf,
            },
            indent=2
        ),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"PROJECT cell: {CELL_PROJECT}")
    print(f"DATE cell: {CELL_DATE} = {ws[CELL_DATE].value!r}  format={ws[CELL_DATE].number_format!r}")
    print(f"ESTIMATOR cell: {CELL_ESTIMATOR} = {args.estimator}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int_money(args.sf)}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow + comment + ROM_LOG)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"SKIPPED (self-perform SUBS block): {len(skipped_self_perf)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
