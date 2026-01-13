#!/usr/bin/env python3
import argparse
import re
from pathlib import Path
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# ===== LOCKED HEADER CELLS (your confirmed layout) =====
CELL_PROJECT   = "O2"   # merged
CELL_ADDR1     = "O4"   # merged (move address UP)
CELL_CITYSTZIP = "O5"   # merged (move city/state UP)
CELL_DATE      = "T2"   # should display as date (NOT serial)
CELL_ESTIMATOR = "T3"   # RNC

CELL_TOTAL_SF  = "H4"
CELL_COST_SF   = "H6"

# ===== LOCKED TABLE COLS (your confirmed layout) =====
# HEADER_ROW is dynamic; these columns are based on your template scans:
COL_CODE = 1      # A
COL_DESC_A = 3    # C header says DESCRIPTION
COL_DESC_B = 2    # B actually contains descriptions in your sheet
COL_UNIT = 5      # E
COL_LBR  = 7      # G
COL_SUBS = 9      # I (UNIT PRICES → SUBS)

# ===== SELF-PERFORMED CODES (NEVER go in SUBS; always go in LBR) =====
# Add/remove as you want; these are "GC-ish" and rough carpentry defaults.
SELF_PERFORMED_CODES = {
    1000,  # many 01xx/01000 GC lines
    6100,  # rough carpentry
    6200,  # finish carpentry (often self-performed in some shops)
    6300,  # millwork coordination (adjust as needed)
    1710, 1711, 1715, 1730, 1750, 1760, 1700,  # cleanup/closeout bucket lines
}

ROM_FILL = PatternFill(patternType="solid", fgColor="FFFF00")  # yellow

def _norm(v) -> str:
    return str(v or "").strip()

def _norm_u(v) -> str:
    return _norm(v).upper()

def _is_colored(cell) -> bool:
    f = cell.fill
    if f is None:
        return False
    # PatternType indicates some fill usage
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr: str):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def _set_int(cell, n: float | int):
    cell.value = int(round(float(n)))
    cell.number_format = "0"

def _set_date_cell(cell, d: date):
    # Write a *date* (not datetime), force a date number format so Excel doesn't show serial.
    cell.value = d
    cell.number_format = "m/d/yyyy"

def find_header_row(ws, max_rows=200) -> int:
    for r in range(1, max_rows + 1):
        row = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, 35)]
        s = " | ".join(row)
        if "CODE" in s and "DESCRIPTION" in s and "UNIT" in s:
            return r
    raise RuntimeError("Could not find header row (CODE/DESCRIPTION/UNIT).")

def choose_desc_col(ws, header_row: int) -> int:
    # Your template has DESCRIPTION header at col C, but the actual text is in col B.
    # We'll prefer the column that has more real strings below the header.
    def score(col: int) -> int:
        sc = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + 220) + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and len(v.strip()) >= 3:
                sc += 1
        return sc
    return COL_DESC_B if score(COL_DESC_B) >= score(COL_DESC_A) else COL_DESC_A

def parse_kv_lines(path: Path) -> dict[int, float]:
    """
    Accepts:
      15500=99782
      16000 = 105097
      # comments ok
    """
    data: dict[int, float] = {}
    if not path:
        return data
    if not path.exists():
        raise RuntimeError(f"File not found: {path}")
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        m = re.match(r"^\s*(\d+)\s*=\s*([0-9,]+(?:\.[0-9]+)?)\s*$", line)
        if not m:
            continue
        code = int(m.group(1))
        amt = float(m.group(2).replace(",", ""))
        data[code] = amt
    return data

def make_amounts_file(ws, header_row: int, desc_col: int, out_path: Path):
    # Build a starter amounts file from rows that look like legitimate write targets:
    # unit is SUB or LS, not colored, not blank desc, code numeric.
    codes = []
    for r in range(header_row + 1, ws.max_row + 1):
        code_v = ws.cell(r, COL_CODE).value
        try:
            code = int(float(code_v))
        except Exception:
            continue

        unit = _norm_u(ws.cell(r, COL_UNIT).value)
        if unit not in ("SUB", "LS"):
            continue

        desc = ws.cell(r, desc_col).value
        if not isinstance(desc, str) or not desc.strip():
            continue

        # avoid colored section headers / subtotal bands
        if _is_colored(ws.cell(r, COL_SUBS)) or _is_colored(ws.cell(r, COL_LBR)):
            continue

        codes.append(code)

    # unique + sorted
    uniq = sorted(set(codes))

    lines = [
        "# TCL starter ROM file (code=amount).",
        "# Put 0 for unknowns; overrides.txt should contain known bid numbers.",
        "# NOTE: Self-performed codes are forced to LBR (never SUBS): " + ", ".join(map(str, sorted(SELF_PERFORMED_CODES))),
        "",
    ]
    for c in uniq:
        lines.append(f"{c}=0")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"WROTE STARTER AMOUNTS: {out_path}")
    print(f"CODES: {len(uniq)}")

def find_write_row(ws, header_row: int, desc_col: int, code: int) -> int | None:
    """
    Pick the first non-colored row for this code where UNIT is SUB or LS
    and the target cell (SUBS/LBR) is not colored.
    """
    want_units = {"SUB", "LS"}
    for r in range(header_row + 1, ws.max_row + 1):
        code_v = ws.cell(r, COL_CODE).value
        try:
            code_i = int(float(code_v))
        except Exception:
            continue
        if code_i != int(code):
            continue

        unit = _norm_u(ws.cell(r, COL_UNIT).value)
        if unit not in want_units:
            continue

        # skip blank descriptions
        dv = ws.cell(r, desc_col).value
        if not (isinstance(dv, str) and dv.strip()):
            continue

        # Avoid writing onto colored “band/section total” rows.
        # Check both possible write targets.
        if _is_colored(ws.cell(r, COL_SUBS)) or _is_colored(ws.cell(r, COL_LBR)):
            continue

        return r
    return None

def write_headers(ws, project: str, addr1: str, citystzip: str, estimator: str):
    _top_left_of_merge(ws, CELL_PROJECT).value = project
    _top_left_of_merge(ws, CELL_ADDR1).value = addr1
    _top_left_of_merge(ws, CELL_CITYSTZIP).value = citystzip
    _top_left_of_merge(ws, CELL_ESTIMATOR).value = estimator

def write_date(ws, d: date):
    c = _top_left_of_merge(ws, CELL_DATE)
    _set_date_cell(c, d)

def write_sf_and_costsf(ws, total_sf: float):
    # H4 is sometimes "colored" in your template - we still write it (you confirmed it works)
    h4 = ws[CELL_TOTAL_SF]
    _set_int(h4, total_sf)

    # H6 formula: Cost/SF = SUM(T column totals) / H4
    ws[CELL_COST_SF].value = f"=IFERROR(SUM($T:$T)/{CELL_TOTAL_SF},0)"
    ws[CELL_COST_SF].number_format = "0"  # whole number

def write_amount(ws, row: int, code: int, amount: float, is_rom: bool):
    # Route self-performed codes to LBR; everything else to SUBS
    target_col = COL_LBR if int(code) in SELF_PERFORMED_CODES else COL_SUBS
    cell = ws.cell(row=row, column=target_col)

    _set_int(cell, amount)

    if is_rom:
        cell.fill = ROM_FILL

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)

    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", default=None, help="ROM file (code=amount). ROM writes get yellow highlight.")
    ap.add_argument("--overrides", default=None, help="Known numbers (code=amount). Overrides do NOT get highlight.")
    ap.add_argument("--make_amounts", action="store_true", help="Generate a starter amounts file and exit.")
    ap.add_argument("--estimator", default="RNC")
    ap.add_argument("--date", default=None, help="Optional: set date (YYYY-MM-DD). If omitted, uses today.")

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)
    desc_col = choose_desc_col(ws, header_row)

    amounts_path = Path(args.amounts) if args.amounts else None
    overrides_path = Path(args.overrides) if args.overrides else None

    if args.make_amounts:
        if not amounts_path:
            raise RuntimeError("--make_amounts requires --amounts <path>")
        make_amounts_file(ws, header_row, desc_col, amounts_path)
        return

    # Header writes (LOCKED)
    write_headers(ws, args.project, args.addr1, args.citystzip, args.estimator)

    # Date (always write a *date*, never time)
    d = None
    if args.date:
        try:
            d = datetime.strptime(args.date.strip(), "%Y-%m-%d").date()
        except Exception:
            raise RuntimeError("--date must be YYYY-MM-DD")
    else:
        d = date.today()
    write_date(ws, d)

    # SF + Cost/SF
    write_sf_and_costsf(ws, args.sf)

    # Load files
    rom = parse_kv_lines(amounts_path) if amounts_path else {}
    overrides = parse_kv_lines(overrides_path) if overrides_path else {}

    # Apply ROM first (yellow), then overrides (no highlight)
    rom_writes = []
    override_writes = []

    # ROM writes: only write non-zero
    for code, amt in sorted(rom.items()):
        if float(amt) == 0:
            continue
        row = find_write_row(ws, header_row, desc_col, code)
        if row is None:
            continue
        write_amount(ws, row, code, amt, is_rom=True)
        rom_writes.append({"code": int(code), "row": int(row), "amount": int(round(float(amt))) })

    # Overrides: write non-zero, and overwrite any ROM value if same code
    for code, amt in sorted(overrides.items()):
        if float(amt) == 0:
            continue
        row = find_write_row(ws, header_row, desc_col, code)
        if row is None:
            continue
        write_amount(ws, row, code, amt, is_rom=False)
        override_writes.append({"code": int(code), "row": int(row), "amount": int(round(float(amt))) })

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(
        __import__("json").dumps({"rom_writes": rom_writes, "override_writes": override_writes}, indent=2),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"HEADER_ROW: {header_row}  CODE_COL:{COL_CODE} DESC_COL:{desc_col} UNIT_COL:{COL_UNIT}  SUBS_COL:{COL_SUBS} LBR_COL:{COL_LBR}")
    print(f"PROJECT cell: {CELL_PROJECT}")
    print(f"DATE cell: {CELL_DATE} = {ws[CELL_DATE].value}")
    print(f"ESTIMATOR cell: {CELL_ESTIMATOR} = {args.estimator}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int(round(args.sf))}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
