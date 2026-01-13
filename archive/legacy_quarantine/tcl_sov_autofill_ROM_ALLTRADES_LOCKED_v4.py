#!/usr/bin/env python3
import argparse
import datetime as dt
import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Column positions (your template)
COL_CODE = 1   # A
COL_DESC = 2   # B  (actual descriptions live here)
COL_UNIT = 5   # E
COL_LBR  = 7   # G  (unit price labor)
COL_SUBS = 9   # I  (unit price subs)

CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"

# Yellow highlight used for ROM writes
FILL_ROM = PatternFill(fill_type="solid", fgColor="FFFF00")

# Codes you consider SELF-PERFORMED (locked to LBR column, not SUBS)
# Add/remove as you want.
SELF_PERFORMED_CODES = {
    1000,   # 01000 GC bucket lines often self-performed
    1600,   # 01600 temp/permits/field office etc (often GC)
    1700,   # 01700 cleanup
    6100,   # 06100 rough carpentry / backing / GC carp
    1710, 1711, 1715, 1730, 1750, 1760,  # cleanup related if you want to lock granular
}

def _norm(v):
    return str(v or "").strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found")

def _find_label(ws, label, rmax=40, cmax=40):
    lab = label.strip().upper()
    for r in range(1, rmax + 1):
        for c in range(1, cmax + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == lab:
                return (r, c)
    return None

def _value_cell_right_of_label(ws, label):
    hit = _find_label(ws, label)
    if not hit:
        raise RuntimeError(f"Could not find label '{label}' in top-left area of sheet.")
    r, c = hit
    # value cell is immediately to the right of the label
    return ws.cell(r, c + 1)

def _guard_can_write(cell, label, allow_colored=False):
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"{label}: target cell is merged.")
    if _is_formula(cell):
        raise RuntimeError(f"{label}: target cell is a formula; refusing to overwrite.")
    if (not allow_colored) and _is_colored(cell):
        raise RuntimeError(f"{label}: target cell is colored/protected; refusing to write.")

def write_headers(ws, project, addr1, citystzip):
    # PROJECT -> cell right of "PROJECT"
    proj_cell = _value_cell_right_of_label(ws, "PROJECT")
    # merged-safe: if value cell is inside merge, write top-left of that merge
    proj_addr = proj_cell.coordinate
    _top_left_of_merge(ws, proj_addr).value = project

    # LOCATION -> cell right of "LOCATION" (this is the first address line)
    loc_cell = _value_cell_right_of_label(ws, "LOCATION")
    loc_addr = loc_cell.coordinate
    _top_left_of_merge(ws, loc_addr).value = addr1

    # City/State/Zip goes on the NEXT ROW in same column as location value cell
    city_cell = ws.cell(loc_cell.row + 1, loc_cell.column)
    _top_left_of_merge(ws, city_cell.coordinate).value = citystzip

    # DATE -> cell right of "DATE"
    date_cell = _value_cell_right_of_label(ws, "DATE")
    date_addr = date_cell.coordinate
    tl = _top_left_of_merge(ws, date_addr)

    # Write as a real date (no time) + force a date format so it never shows 46008
    tl.value = dt.date.today()
    if not tl.number_format or tl.number_format.upper() in ("GENERAL", "@"):
        tl.number_format = "m/d/yyyy"

    # ESTIMATOR -> cell right of "ESTIMATOR"
    est_cell = _value_cell_right_of_label(ws, "ESTIMATOR")
    est_addr = est_cell.coordinate
    _top_left_of_merge(ws, est_addr).value = "RNC"

    return {
        "project_cell": _top_left_of_merge(ws, proj_addr).coordinate,
        "location_cell": _top_left_of_merge(ws, loc_addr).coordinate,
        "city_cell": _top_left_of_merge(ws, city_cell.coordinate).coordinate,
        "date_cell": _top_left_of_merge(ws, date_addr).coordinate,
        "estimator_cell": _top_left_of_merge(ws, est_addr).coordinate,
    }

def write_total_sf_and_costsf(ws, total_sf):
    h4 = ws[CELL_TOTAL_SF]
    # You already confirmed H4 can be colored; we allow writing anyway.
    _guard_can_write(h4, CELL_TOTAL_SF, allow_colored=True)

    h4.value = int(round(float(total_sf)))
    h4.number_format = "0"

    # Cost/SF cell: keep formula-based, no decimals
    h6 = ws[CELL_COST_SF]
    # If it's a value in template, convert to formula you want
    # Use Total (column T) / SF like you locked: =IFERROR(SUM($T:$T)/H4,0)
    h6.value = f"=IFERROR(SUM($T:$T)/{CELL_TOTAL_SF},0)"
    h6.number_format = "0"

def _parse_code(s):
    try:
        return int(float(str(s).strip()))
    except Exception:
        return None

def _load_amounts_text(path: Path):
    """
    amounts.txt format:
      16000=105097
      15500=99782
    """
    if not path.exists():
        raise RuntimeError(f"Amounts file not found: {path}")
    txt = path.read_text(encoding="utf-8", errors="ignore").strip()
    if not txt:
        return {}
    out = {}
    for line in txt.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        m = re.match(r"^\s*([0-9]+)\s*=\s*([0-9,\.]+)\s*$", line)
        if not m:
            continue
        code = int(m.group(1))
        amt = float(m.group(2).replace(",", ""))
        out[code] = amt
    return out

def _load_overrides_text(path: Path):
    # same exact format as amounts.txt
    return _load_amounts_text(path)

def _row_is_writeable(ws, row, target_col):
    cell = ws.cell(row=row, column=target_col)
    if isinstance(cell, MergedCell):
        return False
    if _is_formula(cell):
        return False
    if _is_colored(cell):
        return False
    return True

def _pick_target_rows_for_code(ws, header_row, code, target_col):
    """
    For a given CODE, find the best candidate rows to write into.
    We ONLY write into rows where:
      - CODE matches
      - UNIT is present (not blank) OR description is meaningful
      - target cell is NOT colored and NOT a formula and NOT merged
    We skip the banded/colored subtotal/section rows automatically.
    """
    hits = []
    for r in range(header_row + 1, ws.max_row + 1):
        cval = ws.cell(r, COL_CODE).value
        code_i = _parse_code(cval)
        if code_i != code:
            continue

        # skip obvious "separator" rows (no unit and mostly blank)
        unit = _norm(ws.cell(r, COL_UNIT).value)
        desc = _norm(ws.cell(r, COL_DESC).value)
        if unit == "" and desc == "":
            continue

        if not _row_is_writeable(ws, r, target_col):
            continue

        hits.append(r)

    # Strategy:
    # - If there are multiple rows, prefer rows that look like the "main" line item
    #   (UNIT == SUB or LS tends to be the one you want).
    def score(r):
        unit = _norm(ws.cell(r, COL_UNIT).value)
        desc = _norm(ws.cell(r, COL_DESC).value)
        s = 0
        if unit == "SUB":
            s += 50
        if unit == "LS":
            s += 30
        if "ALLOWANCE" in desc:
            s += 10
        # prefer earlier rows within the code block
        s += max(0, 20 - min(20, (r - header_row)))
        return -s, r

    hits.sort(key=score)
    return hits

def write_amount(ws, row, col, amount, highlight=True):
    cell = ws.cell(row=row, column=col)
    cell.value = int(round(float(amount)))
    cell.number_format = "0"
    if highlight:
        cell.fill = FILL_ROM

def find_header_row(ws):
    # Find row where CODE and DESCRIPTION appear
    for r in range(1, 200):
        row = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, 30)]
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    raise RuntimeError("Could not find header row containing CODE + DESCRIPTION")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", required=False, help="amounts.txt (CODE=AMOUNT per line)")
    ap.add_argument("--overrides", required=False, help="overrides.txt (CODE=AMOUNT per line)")

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)

    # Headers + SF + Cost/SF
    coords = write_headers(ws, args.project, args.addr1, args.citystzip)
    write_total_sf_and_costsf(ws, args.sf)

    # Load numbers
    amounts = _load_amounts_text(Path(args.amounts)) if args.amounts else {}
    overrides = _load_overrides_text(Path(args.overrides)) if args.overrides else {}

    # overrides win
    merged = dict(amounts)
    merged.update(overrides)

    rom_writes = []
    override_writes = []

    for code, amt in merged.items():
        if amt is None:
            continue
        if float(amt) == 0:
            continue

        # self-performed -> LBR, otherwise SUBS
        target_col = COL_LBR if int(code) in SELF_PERFORMED_CODES else COL_SUBS

        rows = _pick_target_rows_for_code(ws, header_row, int(code), target_col)
        if not rows:
            continue

        # Write to the best single row (not every duplicate line)
        r = rows[0]
        write_amount(ws, r, target_col, amt, highlight=True)

        rec = {
            "code": int(code),
            "row": int(r),
            "col": int(target_col),
            "amount": int(round(float(amt))),
            "desc": ws.cell(r, COL_DESC).value,
            "self_performed": (int(code) in SELF_PERFORMED_CODES),
        }
        if int(code) in overrides:
            override_writes.append(rec)
        else:
            rom_writes.append(rec)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(json.dumps(
        {
            "header_cells": coords,
            "rom_writes": rom_writes,
            "override_writes": override_writes,
            "self_performed_codes": sorted(list(SELF_PERFORMED_CODES)),
        },
        indent=2
    ), encoding="utf-8")

    print("SUCCESS")
    print(f"HEADER_ROW: {header_row}  CODE_COL:{COL_CODE} DESC_COL:{COL_DESC} UNIT_COL:{COL_UNIT}  SUBS_COL:{COL_SUBS} LBR_COL:{COL_LBR}")
    print(f"PROJECT cell: {coords['project_cell']}")
    print(f"DATE cell: {coords['date_cell']} = {ws[coords['date_cell']].value}")
    print(f"ESTIMATOR cell: {coords['estimator_cell']} = {ws[coords['estimator_cell']].value}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {ws[CELL_TOTAL_SF].value}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
