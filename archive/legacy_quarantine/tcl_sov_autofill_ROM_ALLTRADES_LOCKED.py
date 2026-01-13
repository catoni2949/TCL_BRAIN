#!/usr/bin/env python3
import argparse
import re
from datetime import date as _date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

SHEET = "ESTIMATE (INPUT)"

# ===== Locked header cells (working set) =====
CELL_PROJECT   = "O2"
CELL_ADDR1     = "O4"
CELL_CITYSTZIP = "O5"
CELL_DATE      = "T2"
CELL_ESTIMATOR = "T3"

CELL_TOTAL_SF  = "H4"
CELL_COST_SF   = "H6"

# ===== Table structure (confirmed by your scans) =====
COL_CODE = 1   # A
COL_DESC = 2   # B  (actual descriptions live here)
COL_UNIT = 5   # E
COL_LBR  = 7   # G (UNIT PRICES -> LBR)
COL_MTL  = 8   # H (UNIT PRICES -> MTL)
COL_SUBS = 9   # I (UNIT PRICES -> SUBS)

# ===== Self-performed rules =====
# These codes are self-performed 99% of the time:
SELF_PERFORM_BLOCK_SUBS = {
    1010, 1020, 1030, 1040, 1050, 1060, 1080,
    1000,
    1600, 1610, 1630, 1640, 1680, 1690,
    1710, 1711, 1715, 1730, 1750, 1760,
    1700,
    2090, 2100, 2200, 2900,
    2000,
    3000, 3010, 3300, 3550, 3710,
    5000, 5100, 5500,
    6100, 6200, 6300,
}

# Special cases: self-perform install labor often, but SUBS buyout still normal/allowed.
SPECIAL_ALLOW_SUBS_AND_LBR = {8750, 10500, 10800}

# Highlight for ROM/assumptions
FILL_ROM = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

def norm(v) -> str:
    return str(v or "").strip().upper()

def is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

def is_colored(cell) -> bool:
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but range not found.")

def safe_write(ws, addr, value):
    cell = top_left_of_merge(ws, addr)
    cell.value = value

def guard_can_write(cell, label):
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"{label}: target is merged.")
    # DO NOT block colored for header cells or total SF; TCL uses formatting heavily.
    # But DO block if it's a formula cell when writing a number.
    return

def find_header_row(ws, max_scan=200):
    for r in range(1, max_scan + 1):
        row = [norm(ws.cell(r, c).value) for c in range(1, 40)]
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    return None

def iter_code_rows(ws, header_row):
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, COL_CODE).value
        if code is None:
            continue
        try:
            code_i = int(float(code))
        except Exception:
            continue
        yield r, code_i

def find_write_row(ws, header_row, code):
    """
    Pick the FIRST non-header row for that code that looks like a real line item (has a unit),
    and avoids the bold/colored subtotal bands when possible.
    """
    candidates = []
    for r, c in iter_code_rows(ws, header_row):
        if c != code:
            continue
        unit = norm(ws.cell(r, COL_UNIT).value)
        if not unit:
            continue
        # Prefer non-colored SUBS cell rows (actual writable lines)
        subs_cell = ws.cell(r, COL_SUBS)
        candidates.append((0 if not is_colored(subs_cell) else 1, r))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]

def int_money(x) -> int:
    return int(round(float(x)))

def write_money_cell(ws, row, col, amount, is_rom=False):
    cell = ws.cell(row=row, column=col)
    # guard formula (we don't overwrite formulas)
    if is_formula(cell):
        return False
    cell.value = int_money(amount)
    if is_rom:
        cell.fill = FILL_ROM
    return True

def write_total_sf_and_rate(ws, sf_value):
    # TOTAL SF is an input cell; write integer
    h4 = ws[CELL_TOTAL_SF]
    if is_formula(h4):
        raise RuntimeError(f"{CELL_TOTAL_SF} is a formula; refusing to overwrite.")
    h4.value = int_money(sf_value)

    # COST/SF formula: leave the template formula if it exists, else set it.
    h6 = ws[CELL_COST_SF]
    if not is_formula(h6):
        h6.value = f"=IFERROR(SUM($T:$T)/{CELL_TOTAL_SF},0)"

def set_date_preserve_excel(ws):
    """
    Write an actual Excel date (not serial), and force a date format with no time.
    """
    dcell = ws[CELL_DATE]
    dcell.value = _date.today()
    dcell.number_format = "m/d/yyyy"

def parse_kv_file(path: Path):
    """
    Supports:
      code=12345               (default behavior)
      code.subs=12345
      code.lbr=12345
      code.mtl=12345
    Comments: lines starting with # or //
    """
    data = {}
    if not path:
        return data
    if not path.exists():
        raise RuntimeError(f"File not found: {path}")
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith("//"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not v:
            continue
        # key: code or code.field
        m = re.match(r"^(\d+)(?:\.(subs|lbr|mtl))?$", k, flags=re.IGNORECASE)
        if not m:
            continue
        code = int(m.group(1))
        field = (m.group(2) or "total").lower()
        try:
            amt = float(v)
        except Exception:
            continue
        if code not in data:
            data[code] = {"total": None, "subs": None, "lbr": None, "mtl": None}
        data[code][field] = amt
    return data

def apply_amount(ws, header_row, code, parts, self_lbr_pct, is_rom, rom_writes, override_writes, skipped_self_perf):
    """
    Default behaviors:
      - Normal code: total -> SUBS
      - Self-perform-block: total -> split into LBR/MTL, SUBS blocked unless code is SPECIAL_ALLOW_SUBS_AND_LBR
      - Special allow subs+lbr: total -> SUBS (buyout default); explicit .lbr/.mtl allowed too
    """
    row = find_write_row(ws, header_row, code)
    if row is None:
        return

    allow_subs = (code not in SELF_PERFORM_BLOCK_SUBS) or (code in SPECIAL_ALLOW_SUBS_AND_LBR)

    # explicit parts take precedence
    subs = parts.get("subs")
    lbr  = parts.get("lbr")
    mtl  = parts.get("mtl")
    total = parts.get("total")

    wrote_any = False

    # If explicit subs provided but subs is blocked -> skip + count
    if subs is not None:
        if allow_subs:
            if write_money_cell(ws, row, COL_SUBS, subs, is_rom=is_rom):
                wrote_any = True
        else:
            skipped_self_perf.append(code)

    # Explicit LBR/MTL always allowed
    if lbr is not None:
        if write_money_cell(ws, row, COL_LBR, lbr, is_rom=is_rom):
            wrote_any = True
    if mtl is not None:
        if write_money_cell(ws, row, COL_MTL, mtl, is_rom=is_rom):
            wrote_any = True

    # If we already handled explicit parts, done.
    if (subs is not None) or (lbr is not None) or (mtl is not None):
        if wrote_any:
            (rom_writes if is_rom else override_writes).append({"code": code, "row": row})
        return

    # Default from total
    if total is None:
        return
    if float(total) == 0:
        return

    if (code in SELF_PERFORM_BLOCK_SUBS) and (code not in SPECIAL_ALLOW_SUBS_AND_LBR):
        # self-performed: split into LBR + MTL
        total_i = int_money(total)
        lbr_i = int(round(total_i * float(self_lbr_pct)))
        mtl_i = total_i - lbr_i
        ok1 = write_money_cell(ws, row, COL_LBR, lbr_i, is_rom=is_rom)
        ok2 = write_money_cell(ws, row, COL_MTL, mtl_i, is_rom=is_rom)
        wrote_any = ok1 or ok2
    else:
        # normal / special: default total goes to SUBS
        if allow_subs:
            wrote_any = write_money_cell(ws, row, COL_SUBS, total, is_rom=is_rom)
        else:
            skipped_self_perf.append(code)

    if wrote_any:
        (rom_writes if is_rom else override_writes).append({"code": code, "row": row})

def make_amounts_from_template(template_path: Path, out_path: Path):
    wb = load_workbook(template_path, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'")
    ws = wb[SHEET]
    header_row = find_header_row(ws)
    if not header_row:
        raise RuntimeError("Could not find header row with CODE/DESCRIPTION")

    codes = []
    seen = set()
    for _, code in iter_code_rows(ws, header_row):
        if code not in seen:
            seen.add(code)
            codes.append(code)
    codes.sort()

    lines = []
    for c in codes:
        lines.append(f"{c}=0")
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return len(codes)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)

    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", default=None, help="ROM amounts file (code=amt or code.subs/lbr/mtl=amt)")
    ap.add_argument("--overrides", default=None, help="Overrides file (same syntax). Overrides win over ROM.")

    ap.add_argument("--make_amounts", action="store_true", help="Create starter amounts file (based on template) and exit.")

    ap.add_argument("--estimator", default="RNC")
    ap.add_argument("--self_lbr_pct", type=float, default=0.60, help="For self-perform totals, % to LBR (rest goes to MTL).")

    args = ap.parse_args()

    template = Path(args.template)
    out = Path(args.out)

    # Make starter amounts and exit
    if args.make_amounts:
        if not args.amounts:
            raise RuntimeError("--make_amounts requires --amounts <path>")
        n = make_amounts_from_template(template, Path(args.amounts))
        print(f"WROTE STARTER AMOUNTS: {args.amounts}")
        print(f"CODES: {n}")
        return

    wb = load_workbook(template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)
    if not header_row:
        raise RuntimeError("Could not find header row with CODE/DESCRIPTION")

    # Headers (merged safe)
    safe_write(ws, CELL_PROJECT, args.project)
    safe_write(ws, CELL_ADDR1, args.addr1)
    safe_write(ws, CELL_CITYSTZIP, args.citystzip)

    # Date + estimator
    set_date_preserve_excel(ws)
    safe_write(ws, CELL_ESTIMATOR, args.estimator)

    # SF + cost/sf formula
    write_total_sf_and_rate(ws, args.sf)

    # Load ROM + overrides
    amounts = parse_kv_file(Path(args.amounts)) if args.amounts else {}
    overrides = parse_kv_file(Path(args.overrides)) if args.overrides else {}

    rom_writes = []
    override_writes = []
    skipped_self_perf = []

    # ROM pass
    for code in sorted(amounts.keys()):
        apply_amount(
            ws, header_row, code, amounts[code],
            self_lbr_pct=args.self_lbr_pct,
            is_rom=True,
            rom_writes=rom_writes,
            override_writes=override_writes,
            skipped_self_perf=skipped_self_perf
        )

    # Overrides pass (wins)
    for code in sorted(overrides.keys()):
        apply_amount(
            ws, header_row, code, overrides[code],
            self_lbr_pct=args.self_lbr_pct,
            is_rom=False,
            rom_writes=rom_writes,
            override_writes=override_writes,
            skipped_self_perf=skipped_self_perf
        )

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(
        __import__("json").dumps(
            {
                "rom_writes": rom_writes,
                "override_writes": override_writes,
                "skipped_self_perf_subs": skipped_self_perf,
            },
            indent=2
        ),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"PROJECT cell: {CELL_PROJECT}")
    print(f"DATE cell: {CELL_DATE} = {ws[CELL_DATE].value!r}  format={ws[CELL_DATE].number_format!r}")
    print(f"ESTIMATOR cell: {CELL_ESTIMATOR} = {args.estimator}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int_money(args.sf)}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"SKIPPED (self-perform SUBS block): {len(skipped_self_perf)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
