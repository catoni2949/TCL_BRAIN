#!/usr/bin/env python3
import argparse
import json
import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Header targets (locked)
HDR_PROJECT   = "O2"
HDR_ADDR1     = "O4"
HDR_CITYSTZIP = "O5"
HDR_ESTIMATOR = "T3"   # value next to ESTIMATOR label at S3
HDR_DATE      = "T2"   # value next to DATE label at S2

# SF cells (locked)
CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"
COST_SF_FORMULA = '=IFERROR(SUM($T:$T)/H4,0)'

# Table columns (locked from your scans)
COL_CODE = 1   # A
COL_DESC = 2   # B (descriptions live here)
COL_UNIT = 4   # D (UNIT)
COL_SUBS = 9   # I (first SUBS unit-price column)

# Conservative ROM $/SF by code (edit later if you want; this is the “estimator brain” starter)
# These are placeholders so the sheet is never blank.
ROM_PER_SF_BY_CODE = {
    9100:  10.00,  # GWB/Studs (example)
    15300:  2.00,  # Fire Protection
    15400:  6.00,  # Plumbing
    15500:  8.00,  # HVAC
    16000:  9.00,  # Electrical base
    16500:  2.50,  # Lighting
    16700:  1.25,  # Data/Comms
    16950:  1.00,  # Fire Alarm
    7100:   3.00,  # Waterproofing
    8100:   4.00,  # Doors/frames
    9100:  10.00,  # Drywall bucket
    9900:   2.50,  # Paint
    6400:   2.00,  # Casework (placeholder)
    11000:  3.00,  # Equipment
    # Add more as you like; anything not listed falls back to default_per_sf
}

DEFAULT_PER_SF = 1.00  # for any SUB row not in ROM_PER_SF_BY_CODE

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for rng in ws.merged_cells.ranges:
        if addr in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _write_date_preserve_format(ws):
    cell = ws[HDR_DATE]
    if isinstance(cell, MergedCell):
        cell = _top_left_of_merge(ws, HDR_DATE)
    fmt = cell.number_format  # preserve template format
    cell.value = datetime.date.today()
    cell.number_format = fmt

def _write_headers(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value   = project
    _top_left_of_merge(ws, HDR_ADDR1).value     = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    ws[HDR_ESTIMATOR].value = "RNC"
    _write_date_preserve_format(ws)

def _write_sf(ws, total_sf):
    ws[CELL_TOTAL_SF].value = float(total_sf)   # you confirmed this is input even if “colored”
    if isinstance(ws[CELL_COST_SF], MergedCell):
        raise RuntimeError(f"{CELL_COST_SF} is merged; cannot write formula.")
    ws[CELL_COST_SF].value = COST_SF_FORMULA

def _load_overrides(path: Path) -> dict[int, float]:
    # optional file: lines like 15500=99782
    if not path:
        return {}
    if not path.exists():
        return {}
    out = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        try:
            code = int(float(k.strip()))
            amt = float(str(v).strip().replace(",", ""))
        except Exception:
            continue
        out[code] = amt
    return out

def _find_first_sub_row_for_code(ws, code: int):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, COL_CODE).value
        try:
            ci = int(float(v))
        except Exception:
            continue
        if ci != code:
            continue
        unit = str(ws.cell(r, COL_UNIT).value or "").strip().upper()
        desc = str(ws.cell(r, COL_DESC).value or "").strip()
        if unit != "SUB":
            continue
        if not desc:
            continue
        cell = ws.cell(r, COL_SUBS)
        if isinstance(cell, MergedCell):
            continue
        if _is_formula(cell):
            continue
        return r, desc
    return None, None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--overrides", help="optional overrides file: code=amount per line")
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    overrides = _load_overrides(Path(args.overrides)) if args.overrides else {}

    _write_headers(ws, args.project, args.addr1, args.citystzip)
    _write_sf(ws, args.sf)

    wrote = []
    # Scan all rows; whenever UNIT == SUB, write ROM into the FIRST row of that code (only once per code)
    seen_codes = set()
    for r in range(1, ws.max_row + 1):
        unit = str(ws.cell(r, COL_UNIT).value or "").strip().upper()
        if unit != "SUB":
            continue

        v = ws.cell(r, COL_CODE).value
        try:
            code = int(float(v))
        except Exception:
            continue
        if code in seen_codes:
            continue

        target_row, desc = _find_first_sub_row_for_code(ws, code)
        if not target_row:
            continue

        # amount priority: override > ROM table > default
        if code in overrides:
            amt = overrides[code]
            src = "OVERRIDE"
        else:
            rate = ROM_PER_SF_BY_CODE.get(code, DEFAULT_PER_SF)
            amt = float(rate) * float(args.sf)
            src = f"ROM_${rate:.2f}/SF"

        ws.cell(target_row, COL_SUBS).value = float(amt)
        wrote.append({"code": code, "row": target_row, "subs_col": COL_SUBS, "amount": amt, "src": src, "desc": desc})
        seen_codes.add(code)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(".rom_writes.json")
    sidecar.write_text(json.dumps({"writes": wrote}, indent=2), encoding="utf-8")

    print("SUCCESS")
    print("DATE:", HDR_DATE, "written (template format preserved)")
    print("ESTIMATOR:", HDR_ESTIMATOR, "= RNC")
    print("TOTAL_SF:", CELL_TOTAL_SF, "=", float(args.sf))
    print("COST/SF:", CELL_COST_SF, "=", COST_SF_FORMULA)
    print("ROM WRITES:", len(wrote))
    print("OUTPUT:", str(out))
    print("SIDECAR:", str(sidecar))

if __name__ == "__main__":
    main()
