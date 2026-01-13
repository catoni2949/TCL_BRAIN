#!/usr/bin/env python3
import argparse
import re
from pathlib import Path
from datetime import date
from typing import Dict, Tuple, Optional, List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SHEET = "ESTIMATE (INPUT)"

# ---- HEADER CELLS (LOCKED) ----
CELL_PROJECT   = "O2"
CELL_ADDR1     = "O4"
CELL_CITYSTZIP = "O5"
CELL_DATE      = "T2"
CELL_ESTIMATOR = "T3"

CELL_TOTAL_SF  = "H4"
CELL_COST_SF   = "H6"

# ---- TABLE COLS (LOCKED BY HEADER SCAN) ----
COL_CODE = 1   # A
COL_DESC = 2   # B  (actual description text lives here)
COL_UNIT = 5   # E
COL_LBR  = 7   # G (not used here but printed in debug)
COL_SUBS = 9   # I (UNIT PRICES -> SUBS)

# ---- SELF-PERFORMED / LOCKED CODES (SKIP WRITES) ----
# Add/remove codes here as you like.
SELF_PERFORMED_CODES = {
    6100,   # ROUGH CARPENTRY (and related in your sheet)
    1000,   # many GC lines are 01000 in some templates; keep if you never want ROM there
    1710, 1711, 1715, 1730, 1750, 1760, 1700,  # cleanup-ish examples (adjust to your actual codes)
}

# ---- HIGHLIGHT FOR ROM CELLS ----
ROM_FILL = PatternFill(patternType="solid", fgColor="FFFF00")  # yellow

def norm(s) -> str:
    return str(s or "").strip().upper()

def is_colored(cell) -> bool:
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        # treat any patterned fill as "protected/colored"
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def set_int(cell, value: float | int):
    v = int(round(float(value)))
    cell.value = v
    cell.number_format = "#,##0"

def write_headers(ws, project: str, addr1: str, citystzip: str, estimator: str, leave_date: bool):
    ws[CELL_PROJECT].value = project
    ws[CELL_ADDR1].value = addr1
    ws[CELL_CITYSTZIP].value = citystzip
    ws[CELL_ESTIMATOR].value = estimator

    # Date: either leave alone, or write a real Excel date (NO TIME) and force display format.
    if not leave_date:
        ws[CELL_DATE].value = date.today()          # date only (no time)
        ws[CELL_DATE].number_format = "m/d/yyyy"    # force display so it never shows 46008

def write_total_sf_and_rate(ws, sf: float):
    set_int(ws[CELL_TOTAL_SF], sf)
    # keep your working formula pattern
    ws[CELL_COST_SF].value = "=IFERROR(SUM($T:$T)/H4,0)"

def load_amounts_txt(path: Path) -> Dict[int, float]:
    """
    amounts.txt format:
      15500=99782
      16000=105097
    Blank lines and # comments ok.
    """
    if not path.exists():
        raise RuntimeError(f"Amounts file not found: {path}")
    out: Dict[int, float] = {}
    for raw in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not k:
            continue
        code = int(float(k))
        amt = float(v) if v else 0.0
        out[code] = amt
    return out

def load_overrides_txt(path: Path) -> Dict[int, float]:
    # same format as amounts.txt
    return load_amounts_txt(path)

def find_header_row(ws) -> int:
    # find row containing CODE + DESCRIPTION
    for r in range(1, 250):
        row = [norm(ws.cell(r, c).value) for c in range(1, 40)]
        joined = " | ".join(row)
        if "CODE" in joined and "DESCRIPTION" in joined:
            return r
    raise RuntimeError("Could not find header row containing CODE and DESCRIPTION.")

def row_matches_code(ws, r: int, code: int) -> bool:
    v = ws.cell(r, COL_CODE).value
    try:
        return int(float(v)) == int(code)
    except Exception:
        return False

def is_writeable_trade_row(ws, r: int) -> bool:
    """
    Only write to rows that look like an actual line item:
    - UNIT column is SUB or LS
    - SUBS cell (I) is not colored
    - Description exists in column B
    """
    unit = norm(ws.cell(r, COL_UNIT).value)
    desc = ws.cell(r, COL_DESC).value
    if unit not in ("SUB", "LS"):
        return False
    if not isinstance(desc, str) or not desc.strip():
        return False
    subs_cell = ws.cell(r, COL_SUBS)
    if is_colored(subs_cell):
        return False
    return True

def pick_best_row_for_code(ws, header_row: int, code: int) -> Optional[int]:
    """
    Prefer first writeable row that matches code.
    If none writeable, return None.
    """
    for r in range(header_row + 1, ws.max_row + 1):
        if not row_matches_code(ws, r, code):
            continue
        if is_writeable_trade_row(ws, r):
            return r
    return None

def write_amount(ws, r: int, code: int, amt: float, is_rom: bool):
    cell = ws.cell(r, COL_SUBS)
    set_int(cell, amt)
    if is_rom:
        cell.fill = ROM_FILL

def make_starter_amounts(ws, header_row: int) -> List[int]:
    """
    Returns list of unique codes that have at least one writeable row (SUB/LS, not colored).
    """
    codes = set()
    for r in range(header_row + 1, ws.max_row + 1):
        if not is_writeable_trade_row(ws, r):
            continue
        v = ws.cell(r, COL_CODE).value
        try:
            code = int(float(v))
        except Exception:
            continue
        codes.add(code)
    return sorted(codes)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--estimator", default="RNC")
    ap.add_argument("--leave_date", action="store_true", help="Do not touch the DATE cell at all.")

    ap.add_argument("--amounts", help="amounts.txt (code=amount).")
    ap.add_argument("--overrides", help="overrides.txt (code=amount). Overrides win over amounts/ROM.")
    ap.add_argument("--make_amounts", action="store_true", help="Generate starter amounts.txt and exit (writes 0s).")

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)

    # A) Headers + SF
    write_headers(ws, args.project, args.addr1, args.citystzip, args.estimator, leave_date=args.leave_date)
    write_total_sf_and_rate(ws, args.sf)

    # Starter file generator
    if args.make_amounts:
        codes = make_starter_amounts(ws, header_row)
        outp = Path(args.amounts or "amounts.txt")
        lines = []
        for c in codes:
            if c in SELF_PERFORMED_CODES:
                lines.append(f"# {c} (self-performed locked)")
                lines.append(f"{c}=0")
            else:
                lines.append(f"{c}=0")
        outp.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"WROTE STARTER AMOUNTS: {outp}")
        print(f"CODES: {len(codes)}")
        return

    # Load inputs
    amounts: Dict[int, float] = load_amounts_txt(Path(args.amounts)) if args.amounts else {}
    overrides: Dict[int, float] = load_overrides_txt(Path(args.overrides)) if args.overrides else {}

    rom_writes = []
    override_writes = []

    # B) Apply amounts as ROM (yellow), but SKIP self-performed and skip zeros
    for code, amt in sorted(amounts.items()):
        if int(code) in SELF_PERFORMED_CODES:
            continue
        if float(amt) == 0:
            continue
        row = pick_best_row_for_code(ws, header_row, int(code))
        if row is None:
            continue
        write_amount(ws, row, int(code), amt, is_rom=True)
        rom_writes.append({"code": int(code), "row": int(row), "amount": int(round(float(amt))) })

    # C) Apply overrides (not yellow), win over ROM
    for code, amt in sorted(overrides.items()):
        if int(code) in SELF_PERFORMED_CODES:
            continue
        if float(amt) == 0:
            continue
        row = pick_best_row_for_code(ws, header_row, int(code))
        if row is None:
            continue
        write_amount(ws, row, int(code), amt, is_rom=False)
        override_writes.append({"code": int(code), "row": int(row), "amount": int(round(float(amt))) })

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(out.suffix + ".rom_writes.json")
    sidecar.write_text(
        __import__("json").dumps({"rom_writes": rom_writes, "override_writes": override_writes}, indent=2),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"HEADER_ROW: {header_row}  CODE_COL:{COL_CODE} DESC_COL:{COL_DESC} UNIT_COL:{COL_UNIT}  SUBS_COL:{COL_SUBS} LBR_COL:{COL_LBR}")
    print(f"PROJECT cell: {CELL_PROJECT}")
    print(f"DATE cell: {CELL_DATE} = {ws[CELL_DATE].value!r}  format={ws[CELL_DATE].number_format!r}")
    print(f"ESTIMATOR cell: {CELL_ESTIMATOR} = {args.estimator}")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int(round(args.sf))}")
    print(f"COST/SF:  {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow)")
    print(f"OVERRIDES:  {len(override_writes)}")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
