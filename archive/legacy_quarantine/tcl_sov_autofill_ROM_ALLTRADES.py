#!/usr/bin/env python3
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Locked columns from your scans
COL_CODE = 1   # A
COL_DESC = 2   # B
COL_UNIT = 5   # E (shows "SUB")
COL_SUBS = 9   # I (first SUBS column)

CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"

ROM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow

def _norm(v): return str(v or "").strip().upper()
def _is_formula(cell): return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None: return False
    if getattr(f, "patternType", None) not in (None, "none"): return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"): return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    return c

def find_label_cell(ws, label, max_r=40, max_c=40):
    target = _norm(label)
    for r in range(1, min(max_r, ws.max_row) + 1):
        for c in range(1, min(max_c, ws.max_column) + 1):
            if _norm(ws.cell(r, c).value) == target:
                return (r, c)
    raise RuntimeError(f"Could not find label '{label}' in top-left area of sheet.")

def write_right_of_label(ws, label, value, row_offset=0):
    r, c = find_label_cell(ws, label)
    tr = r + row_offset
    tc = c + 1
    addr = ws.cell(tr, tc).coordinate
    _top_left_of_merge(ws, addr).value = value
    return addr

def write_headers(ws, project, addr1, citystzip):
    # PROJECT goes right of the PROJECT label
    write_right_of_label(ws, "PROJECT", project)

    # LOCATION label: addr1 on same row, city/state/zip one row below (right of label)
    write_right_of_label(ws, "LOCATION", addr1, row_offset=0)
    write_right_of_label(ws, "LOCATION", citystzip, row_offset=1)

    # Estimator always RNC (right of ESTIMATOR label)
    write_right_of_label(ws, "ESTIMATOR", "RNC")

    # DATE: do NOT touch it (you said leave date alone)

def write_sf(ws, total_sf):
    # You confirmed: Total SF in H4, formula in H6
    ws[CELL_TOTAL_SF].value = float(total_sf)
    ws[CELL_COST_SF].value = "=IFERROR(SUM($T:$T)/H4,0)"

def first_writeable_row_for_code(ws, code_int):
    for r in range(1, ws.max_row + 1):
        code = ws.cell(r, COL_CODE).value
        try:
            code_i = int(float(code))
        except:
            continue
        if code_i != code_int:
            continue

        if _norm(ws.cell(r, COL_UNIT).value) != "SUB":
            continue

        # never touch colored rows (use DESCRIPTION cell as the guard)
        if _is_colored(ws.cell(r, COL_DESC)):
            continue

        subs_cell = ws.cell(r, COL_SUBS)
        if isinstance(subs_cell, MergedCell):
            continue
        if _is_formula(subs_cell):
            continue

        return r
    return None

def rom_amount_for_code(code_int, sf):
    div = code_int // 1000
    defaults = {
        2:  (8.0,   2500),
        6:  (2.0,   1500),   # rough carpentry often self-perform; conservative placeholder only
        8:  (4.0,   2500),
        9:  (9.0,   5000),
        12: (3.5,   2500),
        15: (12.0,  8000),
        16: (10.0,  8000),
    }
    overrides = {
        9100:  (10.0, 10000),
        15300: (6.0,  15000),
        15500: (14.0, 20000),
        16000: (12.0, 20000),
    }
    rate, minv = overrides.get(code_int, defaults.get(div, (5.0, 2500)))
    amt = max(minv, rate * float(sf))
    return round(amt / 50.0) * 50.0

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--rom_only_missing", action="store_true",
                    help="Only fill blank/0 SUBS cells; don't overwrite existing numbers.")
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    write_headers(ws, args.project, args.addr1, args.citystzip)
    write_sf(ws, args.sf)

    codes = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, COL_CODE).value
        try:
            codes.add(int(float(v)))
        except:
            continue

    wrote = []
    for code_int in sorted(codes):
        row = first_writeable_row_for_code(ws, code_int)
        if row is None:
            continue

        subs_cell = ws.cell(row, COL_SUBS)
        existing = subs_cell.value
        if args.rom_only_missing and existing not in (None, "", 0, 0.0):
            continue

        amt = rom_amount_for_code(code_int, args.sf)
        subs_cell.value = float(amt)
        subs_cell.fill = ROM_FILL
        wrote.append((code_int, row, amt, ws.cell(row, COL_DESC).value))

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print("SUCCESS")
    print("ROM WRITES:", len(wrote))
    print("OUTPUT:", str(out))

if __name__ == "__main__":
    main()
