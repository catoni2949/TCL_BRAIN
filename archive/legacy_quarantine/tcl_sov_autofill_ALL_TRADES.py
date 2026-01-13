#!/usr/bin/env python3
import argparse
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# LOCKED ROW MAP (ADD MORE HERE LATER)
ROW_MAP = {
    16000: 210,  # ELECTRICAL BASE (NO FA)
    15500: 202,  # HVAC / MECH
}

COL_SUBS = 9  # Column I

HDR_PROJECT   = "O3"
HDR_ADDR1     = "O5"
HDR_CITYSTZIP = "O6"
HDR_DATE      = "T2"
HDR_ESTIMATOR = "U3"

CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"

def top_left(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(r.min_row, r.min_col)
    raise RuntimeError(f"Merged cell {addr} not resolvable")

def load_amounts(path):
    amounts = {}
    for line in Path(path).read_text().splitlines():
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        amounts[int(k.strip())] = float(v.strip())
    if not amounts:
        raise RuntimeError("amounts.json is empty")
    return amounts

def guarded_write(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"Row {row} col {col} is merged")
    if isinstance(cell.value, str) and cell.value.startswith("="):
        raise RuntimeError(f"Row {row} col {col} is formula")
    cell.value = float(value)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--amounts", required=True)
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True)
    ws = wb[SHEET]

    # Headers
    top_left(ws, HDR_PROJECT).value = args.project
    top_left(ws, HDR_ADDR1).value = args.addr1
    top_left(ws, HDR_CITYSTZIP).value = args.citystzip
    top_left(ws, HDR_DATE).value = date.today().isoformat()
    top_left(ws, HDR_ESTIMATOR).value = "RNC"

    # SF + COST/SF
    top_left(ws, CELL_TOTAL_SF).value = float(args.sf)
    top_left(ws, CELL_COST_SF).value = "=IFERROR(SUM($T:$T)/H4,0)"

    amounts = load_amounts(args.amounts)

    wrote = 0
    for code, amount in amounts.items():
        if code not in ROW_MAP:
            raise RuntimeError(f"CODE {code} not mapped in ROW_MAP")
        row = ROW_MAP[code]
        guarded_write(ws, row, COL_SUBS, amount)
        wrote += 1
        print(f"WROTE CODE {code} → ROW {row} = {amount}")

    if wrote == 0:
        raise RuntimeError("ZERO WRITES — aborting")

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print("SUCCESS →", out)

if __name__ == "__main__":
    main()
