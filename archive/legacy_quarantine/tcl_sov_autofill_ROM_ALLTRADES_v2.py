#!/usr/bin/env python3
import argparse, json, re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

SHEET = "ESTIMATE (INPUT)"

# Header cells (locked/verified in your template)
HDR_PROJECT   = "O3"
HDR_ADDR1     = "O4"
HDR_CITYSTZIP = "O5"
HDR_DATE      = "T2"     # leave format alone; write datetime
HDR_ESTIMATOR = "T3"     # RNC

# SF cells (locked/verified)
CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"

# Unit prices block (locked/verified from scans)
COL_CODE = 1   # A
COL_DESC = 2   # B  (actual descriptions live here)
COL_UNIT = 5   # E  (contains 'SUB', 'LS', etc. for the unit-prices area)
COL_LBR  = 7   # G  (LBR)
COL_SUBS = 9   # I  (SUBS)

ROM_FILL = PatternFill(patternType="solid", fgColor="FFF2CC")  # light yellow

def _norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for rng in ws.merged_cells.ranges:
        if addr in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def write_headers(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value = project
    _top_left_of_merge(ws, HDR_ADDR1).value = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip

    # leave Excel display formatting alone: write a real date-time value
    _top_left_of_merge(ws, HDR_DATE).value = datetime.now()

    _top_left_of_merge(ws, HDR_ESTIMATOR).value = "RNC"

def guard_can_write(cell, label, allow_color=False):
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"{label}: target cell is merged.")
    if _is_formula(cell):
        raise RuntimeError(f"{label}: target cell is a formula.")
    if (not allow_color) and _is_colored(cell):
        raise RuntimeError(f"{label}: target cell is colored/protected.")

def write_total_sf_and_rate(ws, total_sf):
    h4 = ws[CELL_TOTAL_SF]
    # H4 may look "colored" depending on template; allow it (you confirmed it's input)
    guard_can_write(h4, CELL_TOTAL_SF, allow_color=True)
    h4.value = float(total_sf)

    h6 = ws[CELL_COST_SF]
    # keep H6 as formula always
    h6.value = "=IFERROR(SUM($T:$T)/H4,0)"

def iter_code_rows(ws, header_row):
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, COL_CODE).value
        try:
            code_i = int(float(code))
        except Exception:
            continue
        desc = ws.cell(r, COL_DESC).value
        unit = ws.cell(r, COL_UNIT).value
        yield r, code_i, (desc or ""), (unit or "")

def find_header_row(ws):
    # header row has CODE + DESCRIPTION somewhere near the top
    for r in range(1, 200):
        row = " | ".join(_norm(ws.cell(r, c).value) for c in range(1, 30))
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    raise RuntimeError("Could not find header row with CODE + DESCRIPTION")

def write_rom_cell(ws, row, col, amount, highlight=True):
    cell = ws.cell(row=row, column=col)
    guard_can_write(cell, f"ROW {row} COL {col}", allow_color=False)
    cell.value = float(amount)
    if highlight:
        cell.fill = ROM_FILL

def apply_overrides(overrides_path):
    if not overrides_path:
        return {}
    p = Path(overrides_path)
    if not p.exists():
        raise RuntimeError(f"Overrides file not found: {overrides_path}")
    overrides = {}
    for line in p.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip()
        if not k or not v:
            continue
        overrides[int(k)] = float(v)
    return overrides

def load_amounts(amounts_path):
    p = Path(amounts_path)
    if not p.exists():
        raise RuntimeError(f"amounts.json not found: {amounts_path}")
    data = json.loads(p.read_text(encoding="utf-8", errors="ignore"))
    # supports either {"15500": 123} or {"items":[{"code":15500,"amount":123}]}
    out = {}
    if isinstance(data, dict) and "items" in data and isinstance(data["items"], list):
        for it in data["items"]:
            try:
                out[int(it["code"])] = float(it["amount"])
            except Exception:
                pass
        return out
    if isinstance(data, dict):
        for k, v in data.items():
            try:
                out[int(k)] = float(v)
            except Exception:
                pass
        return out
    raise RuntimeError("amounts.json format not recognized")

def try_autopull_from_bids(bids_parsed_path):
    """
    D) Optional: if bids_parsed.json has totals, use them for ELECTRICAL/HVAC.
    If parser didn't extract totals (None), this does nothing.
    """
    if not bids_parsed_path:
        return {}
    p = Path(bids_parsed_path)
    if not p.exists():
        return {}
    j = json.loads(p.read_text(encoding="utf-8", errors="ignore"))
    bids = j.get("bids") or j.get("items") or j.get("parsed_bids") or []
    if not isinstance(bids, list):
        return {}

    pulled = {}
    for b in bids:
        total = b.get("total")
        if total is None:
            continue
        fname = _norm(b.get("file") or "")
        text = " ".join([fname, _norm(b.get("bidder") or ""), _norm(b.get("project_raw") or "")])

        # crude but safe: only two targets for now
        if "ELECT" in text or "KIRBY" in text or "HORIZON" in text:
            pulled[16000] = float(total)
        if "MECH" in text or "HVAC" in text or "ATLAS" in text:
            pulled[15500] = float(total)

    return pulled

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--amounts", required=True, help="amounts.json (starter ROM map)")
    ap.add_argument("--overrides", default=None, help="optional overrides.txt CODE=AMOUNT lines")
    ap.add_argument("--bids_parsed", default=None, help="optional outputs/bids_parsed.json")

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header = find_header_row(ws)
    amounts = load_amounts(args.amounts)
    overrides = apply_overrides(args.overrides)

    # D) optional bid pull (only if totals exist)
    bid_pulled = try_autopull_from_bids(args.bids_parsed)

    # Header + SF
    write_headers(ws, args.project, args.addr1, args.citystzip)
    write_total_sf_and_rate(ws, args.sf)

    wrote = []
    warnings = []

    # Build an index of rows per code (only rows in the unit price section)
    rows_by_code = {}
    for r, code_i, desc, unit in iter_code_rows(ws, header):
        rows_by_code.setdefault(code_i, []).append((r, desc, unit))

    # C) Self-performed rough carpentry:
    # If code 06100 exists, and any row desc contains 'ROUGH CARPENTRY', write to LABOR not SUBS.
    # Conservative ROM rate (change later): $3.50/SF
    ROUGH_CARP_CODE = 6100
    rough_amt = round(float(args.sf) * 3.50, 2)

    # Now write all standard ROM amounts (SUB trades) + special handling for rough carp
    for code, base_amt in amounts.items():
        amt = overrides.get(code, None)
        if amt is None:
            amt = bid_pulled.get(code, None)
        if amt is None:
            amt = base_amt

        if code not in rows_by_code:
            continue

        # Special: rough carp → LABOR col
        if code == ROUGH_CARP_CODE:
            amt = overrides.get(code, bid_pulled.get(code, rough_amt))
            wrote_any = False
            for (r, desc, unit) in rows_by_code[code]:
                if "ROUGH CARPENTRY" not in _norm(desc):
                    continue
                # do NOT write to subtotal/blank rows (those have missing unit/qty patterns)
                if not _norm(unit):
                    continue
                try:
                    write_rom_cell(ws, r, COL_LBR, amt, highlight=True)
                    wrote.append({"code": code, "row": r, "col": "LBR", "amount": amt, "desc": str(desc)})
                    wrote_any = True
                    break  # write first real rough carp row only
                except Exception as e:
                    warnings.append({"code": code, "row": r, "warn": str(e)})
            if not wrote_any:
                warnings.append({"code": code, "warn": "ROUGH CARP row not found"})
            continue

        # Normal: SUB trades → SUBS col
        wrote_one = False
        for (r, desc, unit) in rows_by_code[code]:
            # only write rows that look like real line items (unit present)
            if not _norm(unit):
                continue
            try:
                write_rom_cell(ws, r, COL_SUBS, amt, highlight=True)
                wrote.append({"code": code, "row": r, "col": "SUBS", "amount": amt, "desc": str(desc)})
                wrote_one = True
                break  # one write per code (first valid row)
            except Exception as e:
                warnings.append({"code": code, "row": r, "warn": str(e)})
        if not wrote_one:
            # don't spam warnings for codes that exist but all rows blocked; just note once
            warnings.append({"code": code, "warn": "no_writable_row_found"})

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = out.with_suffix(".rom_writes.json")
    sidecar.write_text(json.dumps({"wrote": wrote, "warnings": warnings}, indent=2), encoding="utf-8")

    print("SUCCESS")
    print(f"HEADER_ROW: {header} CODE_COL: {COL_CODE} DESC_COL_USED: {COL_DESC} UNIT_COL: {COL_UNIT} SUBS_COL: {COL_SUBS} LBR_COL: {COL_LBR}")
    print(f"DATE: {HDR_DATE} written (template format preserved)")
    print(f"ESTIMATOR: {HDR_ESTIMATOR} = RNC")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {float(args.sf)}")
    print(f"COST/SF: {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"ROM WRITES: {len(wrote)}  (yellow highlighted)")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
