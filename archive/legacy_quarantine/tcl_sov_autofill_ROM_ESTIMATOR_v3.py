#!/usr/bin/env python3
import argparse
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

SHEET = "ESTIMATE (INPUT)"

# Header placement (locked)
HDR_PROJECT   = "O3"
HDR_ADDR1     = "O4"   # address goes UP one row
HDR_CITYSTZIP = "O5"   # city/state/zip goes UP one row
HDR_ESTIMATOR = "T3"   # write RNC here (DATE stays untouched)

# SF + Cost/SF
CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"
COST_SF_FORMULA = "=IFERROR(SUM($T:$T)/H4,0)"

# ROM styling
ROM_YELLOW = PatternFill(patternType="solid", fgColor="FFFF00")

def _norm(v):
    return str(v or "").strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but merge range not found.")

def _is_rom_cell(cell):
    # treat our own ROMs as replaceable
    if cell is None:
        return False
    try:
        fg = getattr(cell.fill, "fgColor", None)
        rgb = getattr(fg, "rgb", None) if fg else None
        if rgb and rgb.upper().endswith("FFFF00"):
            return True
    except Exception:
        pass
    if cell.comment and "ROM" in (cell.comment.text or "").upper():
        return True
    return False

def _safe_int_dollars(x):
    # whole dollars, no decimals anywhere
    return int(round(float(x)))

def _round_to_nearest(x, base=1000):
    x = int(round(x))
    return int(base * round(x / base))

def find_header_row(ws, max_rows=200, max_cols=80):
    for r in range(1, max_rows + 1):
        row = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, max_cols + 1)]
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    return None

def find_col_exact(ws, header_row, label, max_cols=80):
    label = label.strip().upper()
    for c in range(1, max_cols + 1):
        v = str(ws.cell(header_row, c).value or "").strip().upper()
        if v == label:
            return c
    return None

def find_first_subs_col(ws, header_row, max_cols=80):
    # Choose the first "SUBS" in the UNIT PRICES block (left side)
    subs_cols = []
    for c in range(1, max_cols + 1):
        v = str(ws.cell(header_row, c).value or "").strip().upper()
        if v == "SUBS":
            subs_cols.append(c)
    if not subs_cols:
        return None
    # In your template, the UNIT PRICES SUBS is the LEFT one (col 9). Take the smallest.
    return min(subs_cols)

def load_overrides(path: Path):
    """
    overrides.txt format:
      16000=105097
      15500=99782
    """
    if not path:
        return {}
    if not path.exists():
        raise RuntimeError(f"Overrides file not found: {path}")
    out = {}
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().replace(",", "")
        if not k or not v:
            continue
        try:
            code = int(k)
            amt = float(v)
        except Exception:
            continue
        out[code] = amt
    return out

def write_headers(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value = project
    _top_left_of_merge(ws, HDR_ADDR1).value = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    # DATE: DO NOT TOUCH
    _top_left_of_merge(ws, HDR_ESTIMATOR).value = "RNC"

def write_sf_and_cost(ws, sf):
    h4 = _top_left_of_merge(ws, CELL_TOTAL_SF)
    if isinstance(h4, MergedCell):
        raise RuntimeError(f"{CELL_TOTAL_SF} is merged; cannot write.")
    # allow writing even if formatted/colored; just don't overwrite formulas
    if _is_formula(h4):
        raise RuntimeError(f"{CELL_TOTAL_SF} is a formula; refusing to overwrite.")
    h4.value = _safe_int_dollars(sf)
    h4.number_format = "0"

    h6 = _top_left_of_merge(ws, CELL_COST_SF)
    if isinstance(h6, MergedCell):
        raise RuntimeError(f"{CELL_COST_SF} is merged; cannot write.")
    # Cost/SF must be formula
    h6.value = COST_SF_FORMULA
    h6.number_format = "0"

def is_self_performed(code: int, desc: str):
    d = _norm(desc)
    if 1000 <= code <= 1999:
        return True
    if code == 6100:
        return True
    if any(k in d for k in ["GENERAL CONDITIONS", "TEMP PROTECTION", "BARRICADE", "CLEANUP", "CLOSEOUT", "GC "]):
        return True
    return False

# ROM $/SF rules (conservative midpoints, rounded to nearest $1,000)
# You can tune these later; no more "amounts.json" nonsense.
ROM_RULES = {
    2090:  (4, 8),    # demo
    3000:  (6, 12),   # concrete
    5100:  (3, 7),    # structural steel
    5500:  (2, 6),    # metal fab
    6200:  (8, 16),   # finish carp / trims
    6300:  (10, 20),  # millwork
    6400:  (10, 22),  # casework
    7100:  (4, 10),   # waterproofing
    7200:  (2, 5),    # insulation
    7300:  (4, 10),   # roofing
    7900:  (1, 3),    # caulking
    8100:  (4, 10),   # metal doors/frames
    8200:  (2, 6),    # wood doors
    8400:  (3, 8),    # storefront/entrances
    8600:  (3, 8),    # windows
    8700:  (4, 10),   # finish hardware
    8750:  (4, 10),   # doors/jambs/hdw
    8800:  (4, 10),   # glazing
    8850:  (1, 4),    # relite
    9100:  (10, 16),  # studs/gwb
    9250:  (9, 14),   # drywall
    9300:  (6, 12),   # feature wall allowance
    9500:  (4, 9),    # ACT
    9600:  (6, 14),   # flooring
    9900:  (3, 7),    # paint
    9950:  (1, 4),    # wall coverings
    10500: (1, 4),    # toilet partitions
    10800: (1, 4),    # accessories
    11000: (2, 6),    # equipment
    15300: (2, 6),    # fire protection
    15400: (6, 10),   # plumbing
    15500: (8, 14),   # hvac
    16000: (10, 16),  # electrical base
    16500: (2, 6),    # lighting (if separated)
    16700: (1, 4),    # comm/data
    16400: (1, 4),    # card access
    16950: (1, 4),    # fire alarm + permit
    16900: (1, 4),    # special systems
}

# Prefer specific description keywords for multi-row codes
CODE_DESC_PREFER = {
    16000: ["BASE"],              # pick "ELECTRICAL BASE (NO FA)" if present
    15500: ["HOOD", "MAU", "EF"],  # pick HVAC / HOOD / MAU / EF if present
    15300: ["FIRE PROTECTION"],
}

def pick_target_row(ws, header_row, code_col, desc_col, unit_col, code: int):
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(r, code_col).value
        try:
            c = int(float(raw))
        except Exception:
            continue
        if c != code:
            continue
        desc = ws.cell(r, desc_col).value
        unit = ws.cell(r, unit_col).value
        # only write on rows that look like line items
        u = _norm(unit)
        if u and u not in ("SUB", "LS"):
            continue
        rows.append((r, _norm(desc or "")))

    if not rows:
        return None

    prefers = CODE_DESC_PREFER.get(code, [])
    if prefers:
        for p in prefers:
            for r, d in rows:
                if p in d:
                    return r
    return rows[0][0]

def guarded_write_rom(ws, row, col, amount, note, allow_replace_rom=False):
    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):
        return False, "merged"

    if _is_formula(cell):
        return False, "formula"

    # if existing value is non-empty:
    if cell.value not in (None, "", 0):
        if allow_replace_rom and _is_rom_cell(cell):
            pass
        else:
            return False, "occupied"

    cell.value = _safe_int_dollars(amount)
    cell.number_format = "0"
    cell.fill = ROM_YELLOW

    # comment
    try:
        cell.comment = Comment(note, "TCL-ROM")
    except Exception:
        pass

    return True, "wrote"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)

    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)

    ap.add_argument("--overrides", default=None)  # optional bids file
    ap.add_argument("--allow_self", action="store_true")  # override self-performed lock

    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]

    header_row = find_header_row(ws)
    if not header_row:
        raise RuntimeError("Could not find header row with CODE + DESCRIPTION")

    code_col = find_col_exact(ws, header_row, "CODE")
    desc_col_hdr = find_col_exact(ws, header_row, "DESCRIPTION")
    unit_col = find_col_exact(ws, header_row, "UNIT")
    subs_col = find_first_subs_col(ws, header_row)

    # In your template, real descriptions live in column B, even though header says C.
    # Use the header if it's column 2; otherwise force to column 2.
    desc_col_used = 2

    if not all([code_col, unit_col, subs_col]):
        raise RuntimeError(f"Header parse failed. header_row={header_row} code_col={code_col} unit_col={unit_col} subs_col={subs_col}")

    # headers + SF
    write_headers(ws, args.project, args.addr1, args.citystzip)
    write_sf_and_cost(ws, args.sf)

    overrides = load_overrides(Path(args.overrides)) if args.overrides else {}

    rom_writes = []
    override_writes = []
    skipped = []

    sf = float(args.sf)

    # First apply OVERRIDES (real bids) — but never overwrite real numbers
    for code, amt in overrides.items():
        trow = pick_target_row(ws, header_row, code_col, desc_col_used, unit_col, int(code))
        if not trow:
            skipped.append({"code": int(code), "reason": "no_row"})
            continue

        desc = ws.cell(trow, desc_col_used).value or ""
        if (not args.allow_self) and is_self_performed(int(code), desc):
            skipped.append({"code": int(code), "row": trow, "reason": "self_locked"})
            continue

        ok, why = guarded_write_rom(
            ws, trow, subs_col, amt,
            note=f"BID OVERRIDE — CODE {int(code)} — {str(desc).strip()}",
            allow_replace_rom=True
        )
        if ok:
            override_writes.append({"code": int(code), "row": trow, "amount": _safe_int_dollars(amt), "desc": str(desc).strip()})
        else:
            skipped.append({"code": int(code), "row": trow, "reason": why})

    # Then generate ROMs for remaining rules
    for code, (mn, mx) in sorted(ROM_RULES.items()):
        code = int(code)

        # don't ROM if already overridden
        if code in overrides:
            continue

        trow = pick_target_row(ws, header_row, code_col, desc_col_used, unit_col, code)
        if not trow:
            continue

        desc = ws.cell(trow, desc_col_used).value or ""
        if (not args.allow_self) and is_self_performed(code, desc):
            continue

        midpoint = (mn + mx) / 2.0
        rom = sf * midpoint
        rom = _round_to_nearest(rom, base=1000)

        ok, why = guarded_write_rom(
            ws, trow, subs_col, rom,
            note=f"ROM — no bid yet — ${mn}-{mx}/SF midpoint=${midpoint:.1f} — SF={int(round(sf))}",
            allow_replace_rom=True
        )
        if ok:
            rom_writes.append({"code": code, "row": trow, "amount": int(rom), "desc": str(desc).strip()})

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    sidecar = Path(str(out) + ".rom_writes.json")
    sidecar.write_text(
        __import__("json").dumps(
            {
                "rom_writes": rom_writes,
                "override_writes": override_writes,
                "skipped": skipped,
                "meta": {
                    "template": str(args.template),
                    "output": str(out),
                    "header_row": header_row,
                    "code_col": code_col,
                    "desc_col_used": desc_col_used,
                    "unit_col": unit_col,
                    "subs_col": subs_col,
                    "sf": int(round(sf)),
                    "date_touched": False,
                    "self_performed_locked": (not args.allow_self),
                }
            },
            indent=2
        ),
        encoding="utf-8"
    )

    print("SUCCESS")
    print(f"HEADER_ROW: {header_row} CODE_COL: {code_col} DESC_COL_USED: {desc_col_used} UNIT_COL: {unit_col} SUBS_COL: {subs_col}")
    print("DATE: untouched")
    print(f"ESTIMATOR: {HDR_ESTIMATOR} = RNC")
    print(f"TOTAL_SF: {CELL_TOTAL_SF} = {int(round(sf))}")
    print(f"COST/SF: {CELL_COST_SF} = {ws[CELL_COST_SF].value}")
    print(f"OVERRIDES WRITES: {len(override_writes)}")
    print(f"ROM WRITES: {len(rom_writes)} (yellow highlighted)")
    print(f"OUTPUT: {out}")
    print(f"SIDECAR: {sidecar}")

if __name__ == "__main__":
    main()
