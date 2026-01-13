#!/usr/bin/env python3
import argparse
import re
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell

SHEET_NAME = "ESTIMATE (INPUT)"

# Locked cell addresses from your template (write-safe for merged cells)
HEADER_ADDR = {
    "PROJECT": "O3",
    "ADDR1": "O4",
    "CITYSTZIP": "O5",
    "DATE": "T3",
    "ESTIMATOR": "T5",  # ALWAYS RNC
}

# HARD MAP (electrical + mechanical only)
# Electrical is CODE 16000 row(s)
# Mechanical is HVAC rows = CODE 15500 (your template has 15500, NOT 15000)
TRADE_TARGETS = {
    "ELECTRICAL": {"code": 16000, "must_contain": "ELECTRICAL"},
    "MECHANICAL": {"code": 15500, "must_contain": "HVAC"},
}

def norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().upper()

def is_colored(cell) -> bool:
    f = cell.fill
    if f is None:
        return False
    pt = getattr(f, "patternType", None)
    if pt not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    if fg is not None:
        rgb = getattr(fg, "rgb", None)
        if rgb and rgb not in ("00000000", "FFFFFFFF"):
            return True
    return False

def is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

def write_safe(ws, addr: str, value):
    """Write even if addr points into a merged region (writes to top-left)."""
    cell = ws[addr]
    if isinstance(cell, MergedCell):
        # find the merged range containing this cell, write to its top-left
        for r in ws.merged_cells.ranges:
            if addr in r:
                tl = f"{get_column_letter(r.min_col)}{r.min_row}"
                ws[tl].value = value
                return
        raise RuntimeError(f"Cell {addr} is merged but range not found")
    cell.value = value

def find_header_row(ws, max_scan=200):
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row = [norm(ws.cell(r, c).value) for c in range(1, min(80, ws.max_column) + 1)]
        if "CODE" in row and "DESCRIPTION" in row:
            return r
    raise RuntimeError("Could not find header row containing CODE + DESCRIPTION")

def find_col_exact(ws, header_row, header_text, max_cols=80):
    target = norm(header_text)
    for c in range(1, min(max_cols, ws.max_column) + 1):
        if norm(ws.cell(header_row, c).value) == target:
            return c
    return None

def find_subs_col(ws, header_row):
    # Your template has multiple "SUBS". We want the FIRST one (unit price input area).
    subs_cols = []
    for c in range(1, min(40, ws.max_column) + 1):
        if norm(ws.cell(header_row, c).value) == "SUBS":
            subs_cols.append(c)
    if not subs_cols:
        raise RuntimeError("Could not find any column named SUBS on header row")
    return subs_cols[0]

def find_sf_input_cell(ws):
    """
    Find an SF input cell by locating a label like 'SF' / 'SQUARE FOOT' in top area,
    then picking the nearest empty numeric cell to the right on that row.
    """
    needles = ["SQUARE", "SQUARE FOOT", "SQUARE FOOTAGE", "SF", "SQ FT"]
    for r in range(1, 60):
        for c in range(1, min(60, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            t = norm(v)
            if any(n in t for n in needles):
                # look right for a writable input cell
                for cc in range(c + 1, min(c + 12, ws.max_column) + 1):
                    cell = ws.cell(r, cc)
                    if isinstance(cell, MergedCell):
                        continue
                    if is_formula(cell) or is_colored(cell):
                        continue
                    # prefer blank or 0
                    if cell.value in (None, "", 0):
                        return f"{get_column_letter(cc)}{r}"
    return None  # not fatal

def find_target_rows(ws, header_row, code_col, desc_col, target_code, must_contain):
    hits = []
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        try:
            code_i = int(float(code))
        except Exception:
            continue
        if code_i != target_code:
            continue
        desc = norm(ws.cell(r, desc_col).value)
        if must_contain and must_contain not in desc:
            continue
        hits.append(r)
    return hits

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--mech", type=float, required=True)
    ap.add_argument("--elec", type=float, required=True)
    args = ap.parse_args()

    tpl = Path(args.template)
    if not tpl.exists():
        raise RuntimeError(f"Template not found: {tpl}")

    wb = load_workbook(str(tpl), keep_vba=True, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET_NAME}'. Found: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # Header fill
    write_safe(ws, HEADER_ADDR["PROJECT"], args.project)
    write_safe(ws, HEADER_ADDR["ADDR1"], args.addr1)
    write_safe(ws, HEADER_ADDR["CITYSTZIP"], args.citystzip)
    write_safe(ws, HEADER_ADDR["DATE"], date.today().isoformat())
    write_safe(ws, HEADER_ADDR["ESTIMATOR"], "RNC")

    # SF fill (write to input cell only; never touch formula cells)
    sf_cell = find_sf_input_cell(ws)
    if sf_cell:
        write_safe(ws, sf_cell, float(args.sf))

    # Find table structure
    header_row = find_header_row(ws)
    code_col = find_col_exact(ws, header_row, "CODE")
    desc_col = find_col_exact(ws, header_row, "DESCRIPTION")
    if not code_col or not desc_col:
        raise RuntimeError(f"Could not find CODE/DESCRIPTION columns on header row {header_row}")
    subs_col = find_subs_col(ws, header_row)

    # HARD MAP write
    wrote = []
    trade_amounts = {
        "ELECTRICAL": float(args.elec),
        "MECHANICAL": float(args.mech),
    }

    for trade, spec in TRADE_TARGETS.items():
        target_code = spec["code"]
        must = spec["must_contain"]
        rows = find_target_rows(ws, header_row, code_col, desc_col, target_code, must)
        if not rows:
            raise RuntimeError(f"Could not find target row(s) for {trade}: CODE {target_code} containing '{must}'.")

        # choose first matching row that is writable in SUBS column
        amount = trade_amounts[trade]
        wrote_one = False
        for r in rows:
            cell = ws.cell(r, subs_col)
            if isinstance(cell, MergedCell):
                continue
            if is_formula(cell) or is_colored(cell):
                continue
            cell.value = float(amount)
            wrote.append({"trade": trade, "row": r, "subs_col": subs_col, "amount": amount})
            wrote_one = True
            break

        if not wrote_one:
            raise RuntimeError(f"Found {trade} row(s) but SUBS cell was protected/colored/formula for all of them.")

    if len(wrote) != 2:
        raise RuntimeError(f"Expected 2 writes (MECH+ELEC). Got {len(wrote)}. Aborting.")

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    print("SUCCESS")
    for w in wrote:
        print(f"WROTE {w['trade']} -> row {w['row']} (SUBS col {w['subs_col']}) = {w['amount']}")
    if sf_cell:
        print(f"WROTE SF input -> {sf_cell} = {args.sf}")
    else:
        print("SF label not found; skipped SF input write (template-specific).")
    print("OUTPUT:", out)

if __name__ == "__main__":
    main()
