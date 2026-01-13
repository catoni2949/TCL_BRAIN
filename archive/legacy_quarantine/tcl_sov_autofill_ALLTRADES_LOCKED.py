#!/usr/bin/env python3
import argparse
import re
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SHEET = "ESTIMATE (INPUT)"

# Locked columns (confirmed from your scans)
COL_CODE = 1   # A
COL_DESC = 2   # B
COL_UNIT = 5   # E  (this is where "SUB" lives on your lines)
COL_SUBS = 9   # I  (first SUBS column in UNIT PRICES)

# Locked header cells (merged-safe)
HDR_PROJECT   = "O3"
HDR_ADDR1     = "O4"
HDR_CITYSTZIP = "O5"
HDR_ESTIMATOR = "U3"   # locked now
# DATE: DO NOT TOUCH

CELL_TOTAL_SF = "H4"
CELL_COST_SF  = "H6"
COST_SF_FORMULA = '=IFERROR(SUM($T:$T)/H4,0)'

def _norm(v):
    return str(v or "").strip().upper()

def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")

def _is_colored(cell):
    f = cell.fill
    if f is None:
        return False
    if getattr(f, "patternType", None) not in (None, "none"):
        return True
    fg = getattr(f, "fgColor", None)
    rgb = getattr(fg, "rgb", None) if fg else None
    if rgb and rgb not in ("00000000", "FFFFFFFF"):
        return True
    return False

def _top_left_of_merge(ws, addr):
    c = ws[addr]
    if not isinstance(c, MergedCell):
        return c
    for r in ws.merged_cells.ranges:
        if addr in r:
            return ws.cell(row=r.min_row, column=r.min_col)
    raise RuntimeError(f"Cell {addr} is merged but range not found.")

def _read_amounts(path: Path):
    """
    Accepts:
      - lines like: 16000=105097
      - ignores blank lines and comments starting with #
    """
    out = {}
    txt = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    for line in txt:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        m = re.match(r"^\s*(\d+)\s*=\s*([0-9,]+(?:\.[0-9]+)?)\s*$", line)
        if not m:
            raise RuntimeError(f"Bad line in amounts file: {line!r} (expected CODE=NUMBER)")
        code = int(m.group(1))
        amt = float(m.group(2).replace(",", ""))
        out[code] = amt
    return out

def _find_header_row(ws):
    # You already confirmed header row is 11, but we keep it safe
    for r in range(1, 200):
        s = " ".join(str(ws.cell(r,c).value or "").upper() for c in range(1, 40))
        if "CODE" in s and "DESCRIPTION" in s:
            return r
    raise RuntimeError("Could not find header row with CODE + DESCRIPTION")

def _guard_write(cell, label):
    if isinstance(cell, MergedCell):
        raise RuntimeError(f"{label}: merged cell (refusing).")
    if _is_formula(cell):
        raise RuntimeError(f"{label}: is a formula (refusing).")
    if _is_colored(cell):
        raise RuntimeError(f"{label}: colored/protected (refusing).")

def write_headers(ws, project, addr1, citystzip):
    _top_left_of_merge(ws, HDR_PROJECT).value = project
    _top_left_of_merge(ws, HDR_ADDR1).value = addr1
    _top_left_of_merge(ws, HDR_CITYSTZIP).value = citystzip
    _top_left_of_merge(ws, HDR_ESTIMATOR).value = "RNC"

def write_sf(ws, total_sf):
    # H4 can be colored in your template — allow writing anyway.
    h4 = ws[CELL_TOTAL_SF]
    if isinstance(h4, MergedCell):
        raise RuntimeError("H4 is merged (unexpected).")
    h4.value = float(total_sf)

    # H6 must be a formula — force it.
    h6 = ws[CELL_COST_SF]
    if isinstance(h6, MergedCell):
        raise RuntimeError("H6 is merged (unexpected).")
    h6.value = COST_SF_FORMULA

def find_target_row_for_code(ws, header_row, code):
    """
    Picks the FIRST row for that CODE where:
      - UNIT column == 'SUB'
      - DESCRIPTION is present
    """
    for r in range(header_row+1, ws.max_row+1):
        v = ws.cell(r, COL_CODE).value
        try:
            code_i = int(float(v))
        except Exception:
            continue
        if code_i != int(code):
            continue

        unit = _norm(ws.cell(r, COL_UNIT).value)
        desc = ws.cell(r, COL_DESC).value

        if unit != "SUB":
            continue
        if not (isinstance(desc, str) and desc.strip()):
            continue
        return r, (desc.strip())
    return None, None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--project", required=True)
    ap.add_argument("--addr1", required=True)
    ap.add_argument("--citystzip", required=True)
    ap.add_argument("--sf", type=float, required=True)
    ap.add_argument("--amounts", required=True, help="Path to amounts file lines CODE=AMOUNT")
    ap.add_argument("--make_amounts", action="store_true",
                    help="Generate a starter amounts file (all CODE rows that have UNIT=SUB), with 0s, then exit.")
    args = ap.parse_args()

    wb = load_workbook(args.template, keep_vba=True, data_only=False)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"Missing sheet '{SHEET}'. Found: {wb.sheetnames}")
    ws = wb[SHEET]
    header = _find_header_row(ws)

    if args.make_amounts:
        outp = Path(args.amounts)
        codes = []
        seen = set()
        for r in range(header+1, ws.max_row+1):
            v = ws.cell(r, COL_CODE).value
            try:
                code_i = int(float(v))
            except Exception:
                continue
            unit = _norm(ws.cell(r, COL_UNIT).value)
            desc = ws.cell(r, COL_DESC).value
            if unit != "SUB":
                continue
            if not (isinstance(desc, str) and desc.strip()):
                continue
            if code_i in seen:
                continue
            seen.add(code_i)
            codes.append(code_i)

        codes.sort()
        lines = ["# TCL amounts file", "# format: CODE=AMOUNT", ""]
        for c in codes:
            lines.append(f"{c}=0")
        outp.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print("WROTE STARTER AMOUNTS:", str(outp))
        print("CODES:", len(codes))
        return

    amounts = _read_amounts(Path(args.amounts))

    # Header + SF
    write_headers(ws, args.project, args.addr1, args.citystzip)
    write_sf(ws, args.sf)

    wrote = []
    for code, amt in sorted(amounts.items()):
        if amt == 0:
            continue
        row, desc = find_target_row_for_code(ws, header, code)
        if row is None:
            raise RuntimeError(f"Could not find target row for CODE {code} with UNIT=SUB.")
        cell = ws.cell(row=row, column=COL_SUBS)
        _guard_write(cell, f"SUBS cell for CODE {code} row {row}")
        cell.value = float(amt)
        wrote.append({"code": code, "row": row, "subs_col": COL_SUBS, "amount": amt, "desc": desc})

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print("SUCCESS")
    print("WROTE_COUNT:", len(wrote))
    for w in wrote[:20]:
        print(f"CODE {w['code']} row {w['row']} -> I{w['row']} = {w['amount']}  ({w['desc']})")
    if len(wrote) > 20:
        print("... more writes not shown")
    print("OUTPUT:", str(out))

if __name__ == "__main__":
    main()
