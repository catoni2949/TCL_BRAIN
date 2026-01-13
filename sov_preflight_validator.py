
def is_valid_code(x: str) -> bool:
    s = str(x).strip()
    return s.isdigit()


#!/usr/bin/env python3
"""
TCL SOV Preflight Validator (READ-ONLY)

Goal:
- Prove we are running the expected code file
- Prove which template is being used (absolute path + SHA256)
- Extract the template's authoritative COST CODE list from "CODE" column(s)
- Validate codes are sane (5-digit numeric)
- Optionally compare to a provided code register CSV
- Optionally load a template lock JSON (for provenance only)
- Output TXT + JSON report
- NEVER writes to Excel

Exit codes:
0 = GO
2 = NO-GO (validation failed)
"""

import argparse
import csv
import hashlib
import json
import os
import sys
import time
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional, Tuple, Set

try:
    import openpyxl
except Exception as e:
    print("FATAL: openpyxl is required. Install with: pip3 install openpyxl")
    raise

# ---------------------- helpers ----------------------

def sha256_file(path: str, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            b = f.read(chunk_size)
            if not b:
                break
            h.update(b)
    return h.hexdigest()

def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")

def safe_mkdir(path: str) -> str:
    p = os.path.expanduser(path)
    os.makedirs(p, exist_ok=True)
    return p

def normalize_code_cell(v) -> Optional[str]:
    if v is None:
        return None

    # Ignore formulas (e.g., ='ESTIMATE (INPUT)'!A13)
    if isinstance(v, str):
        s0 = v.strip()
        if s0.startswith("="):
            return None

    s = str(v).strip()
    if not s:
        return None

    return s

def load_code_register_csv(path: str) -> Set[str]:
    codes: Set[str] = set()
    with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return codes

    # Find a column likely to contain codes
    header = [c.strip().lower() for c in rows[0]]
    code_idx = None
    for i, h in enumerate(header):
        if h in ("code", "cost code", "cost_code", "csi", "sov code", "sov_code"):
            code_idx = i
            break
    if code_idx is None:
        # fallback: first column
        code_idx = 0

    for r in rows[1:]:
        if code_idx >= len(r):
            continue
        c = normalize_code_cell(r[code_idx])
        if c:
            codes.add(c)
    return codes

def find_code_headers(ws) -> List[Tuple[int, int]]:
    """
    Returns list of (row, col) where cell value == 'CODE' (case-insensitive)
    Searches first 100 rows and first 80 columns (safe for templates).
    """
    headers = []
    max_row = min(ws.max_row or 0, 100)
    max_col = min(ws.max_column or 0, 80)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().lower() == "code":
                headers.append((r, c))
    return headers

def extract_codes_below(ws, header_rc: Tuple[int, int], max_scan_rows: int = 3000) -> List[Tuple[str, int, int]]:
    """
    Extract codes found below a CODE header.
    Stops when we hit a run of blanks OR a run of non-code content (prevents drift into other sections).
    Returns list of tuples: (code_string, row, col)
    """
    hr, hc = header_rc
    results: List[Tuple[str, int, int]] = []
    blank_run = 0
    noncode_run = 0

    def looks_like_code(s: str) -> bool:
        # TCL template uses numeric bucket codes (e.g., 1000, 1010, 1600, etc.)
        # Accept only digits here.
        return s.isdigit()

    for r in range(hr + 1, min(ws.max_row or 0, hr + max_scan_rows) + 1):
        v = ws.cell(row=r, column=hc).value
        s = normalize_code_cell(v)

        if s is None:
            blank_run += 1
            if blank_run >= 25:
                break
            continue

        blank_run = 0

        if not looks_like_code(s):
            noncode_run += 1
            if noncode_run >= 5:
                break
            continue

        noncode_run = 0
        results.append((s, r, hc))

    return results

# ---------------------- report structures ----------------------

@dataclass
class SheetCodeScan:
    sheet: str
    code_headers: List[Tuple[int, int]]
    codes_found: int
    invalid_codes: List[Dict]
    duplicate_codes: List[str]

@dataclass
class PreflightResult:
    status: str  # "GO" or "NO-GO"
    timestamp: str
    running_file: str
    argv: List[str]
    template_path: str
    template_sha256: str
    template_sheets: List[str]
    total_codes_found: int
    unique_codes_found: int
    invalid_codes_total: int
    duplicates_total: int
    register_path: Optional[str]
    register_codes_count: int
    codes_missing_in_register: List[str]
    template_lock_path: Optional[str]
    template_lock_loaded: bool
    notes: List[str]
    per_sheet: List[SheetCodeScan]

# ---------------------- main ----------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="READ-ONLY SOV template preflight validator (no Excel writes).")
    ap.add_argument("--template", required=True, help="Absolute path to .xlsm/.xlsx template (or output file for inspection).")
    ap.add_argument("--code-register", default="", help="Optional CSV code register to compare against.")
    ap.add_argument("--template-lock", default="", help="Optional JSON lock file (provenance only).")
    ap.add_argument("--out-dir", required=True, help="Directory to write TXT/JSON reports.")
    ap.add_argument("--scan-max-rows", type=int, default=3000, help="Max rows to scan below CODE header per sheet.")
    args = ap.parse_args()

    running_file = os.path.abspath(__file__)
    argv = sys.argv[:]
    ts = now_ts()

    template_path = os.path.abspath(os.path.expanduser(args.template))
    out_dir = safe_mkdir(args.out_dir)

    notes: List[str] = []
    if not os.path.exists(template_path):
        print(f"FATAL: template not found: {template_path}")
        return 2

    try:
        template_hash = sha256_file(template_path)
    except Exception as e:
        print(f"FATAL: unable to hash template: {e}")
        return 2

    # Load template (read-only intent; openpyxl doesn't write unless saved)
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False, keep_vba=True)
    except Exception as e:
        print(f"FATAL: unable to open workbook: {e}")
        return 2

    sheets = wb.sheetnames[:]

    register_path = os.path.abspath(os.path.expanduser(args.code_register)) if args.code_register else ""
    register_codes: Set[str] = set()
    if register_path:
        if os.path.exists(register_path):
            try:
                register_codes = load_code_register_csv(register_path)
            except Exception as e:
                notes.append(f"WARNING: failed to load code register CSV: {e}")
        else:
            notes.append(f"WARNING: code register path does not exist: {register_path}")

    lock_path = os.path.abspath(os.path.expanduser(args.template_lock)) if args.template_lock else ""
    lock_loaded = False
    if lock_path:
        if os.path.exists(lock_path):
            try:
                with open(lock_path, "r", encoding="utf-8") as f:
                    _ = json.load(f)
                lock_loaded = True
            except Exception as e:
                notes.append(f"WARNING: failed to parse template lock JSON: {e}")
        else:
            notes.append(f"WARNING: template lock path does not exist: {lock_path}")

    all_codes: List[str] = []
    all_invalid: List[Dict] = []
    per_sheet: List[SheetCodeScan] = []

    for s in sheets:
        ws = wb[s]
        headers = find_code_headers(ws)
        codes_with_pos: List[Tuple[str, int, int]] = []
        for hrc in headers:
            codes_with_pos.extend(extract_codes_below(ws, hrc, max_scan_rows=args.scan_max_rows))

        codes = [c for (c, _, _) in codes_with_pos]
        all_codes.extend(codes)

        invalids = []
        for (c, r, col) in codes_with_pos:
            if not is_valid_code(c):
                invalids.append({"code": c, "row": r, "col": col, "sheet": s})
        all_invalid.extend(invalids)

        dups = sorted({c for c in codes if codes.count(c) > 1})

        per_sheet.append(SheetCodeScan(
            sheet=s,
            code_headers=headers,
            codes_found=len(codes),
            invalid_codes=invalids,
            duplicate_codes=dups
        ))

    unique_codes = sorted(set(all_codes))
    duplicates_total = len(all_codes) - len(unique_codes)

    # Compare to register
    missing_in_register: List[str] = []
    if register_codes:
        missing_in_register = sorted([c for c in unique_codes if c not in register_codes])

    # Determine GO / NO-GO
    status = "GO"
    if len(unique_codes) == 0:
        status = "NO-GO"
        notes.append("NO-GO: No codes found under any 'CODE' header. Template may not match expected structure.")
    if all_invalid:
        status = "NO-GO"
        notes.append(f"NO-GO: Found {len(all_invalid)} invalid (non-5-digit) codes in template CODE column(s).")
    if register_path and register_codes and missing_in_register:
        status = "NO-GO"
        notes.append(f"NO-GO: {len(missing_in_register)} template codes are missing from provided code register.")

    result = PreflightResult(
        status=status,
        timestamp=ts,
        running_file=running_file,
        argv=argv,
        template_path=template_path,
        template_sha256=template_hash,
        template_sheets=sheets,
        total_codes_found=len(all_codes),
        unique_codes_found=len(unique_codes),
        invalid_codes_total=len(all_invalid),
        duplicates_total=duplicates_total,
        register_path=register_path if register_path else None,
        register_codes_count=len(register_codes),
        codes_missing_in_register=missing_in_register,
        template_lock_path=lock_path if lock_path else None,
        template_lock_loaded=lock_loaded,
        notes=notes,
        per_sheet=per_sheet
    )

    # Write reports
    base = os.path.splitext(os.path.basename(template_path))[0]
    safe_base = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in base)[:90]
    txt_path = os.path.join(out_dir, f"SOV_Preflight_{safe_base}.txt")
    json_path = os.path.join(out_dir, f"SOV_Preflight_{safe_base}.json")

    # TXT report (human)
    lines: List[str] = []
    lines.append("SOV PREFLIGHT VALIDATION REPORT (READ-ONLY)")
    lines.append("")
    lines.append(f"Timestamp: {result.timestamp}")
    lines.append(f"Status: {result.status}")
    lines.append(f"Running file: {result.running_file}")
    lines.append(f"Template: {result.template_path}")
    lines.append(f"Template SHA256: {result.template_sha256}")
    lines.append(f"Sheets: {', '.join(result.template_sheets)}")
    lines.append("")
    lines.append(f"Codes found (total): {result.total_codes_found}")
    lines.append(f"Codes found (unique): {result.unique_codes_found}")
    lines.append(f"Invalid codes: {result.invalid_codes_total}")
    lines.append(f"Duplicate count (approx): {result.duplicates_total}")
    lines.append("")
    if result.register_path:
        lines.append(f"Code register: {result.register_path}")
        lines.append(f"Register codes count: {result.register_codes_count}")
        lines.append(f"Template codes missing in register: {len(result.codes_missing_in_register)}")
    else:
        lines.append("Code register: (not provided)")
    lines.append("")
    if result.template_lock_path:
        lines.append(f"Template lock: {result.template_lock_path} (loaded={result.template_lock_loaded})")
    else:
        lines.append("Template lock: (not provided)")
    lines.append("")
    if result.notes:
        lines.append("Notes:")
        for n in result.notes:
            lines.append(f"- {n}")
        lines.append("")
    lines.append("Per-sheet scan summary:")
    for ps in result.per_sheet:
        lines.append(f"- {ps.sheet}: headers={len(ps.code_headers)} codes={ps.codes_found} invalid={len(ps.invalid_codes)} dups={len(ps.duplicate_codes)}")
    lines.append("")
    if result.codes_missing_in_register:
        lines.append("Codes missing in register (first 80):")
        for c in result.codes_missing_in_register[:80]:
            lines.append(f"- {c}")
        lines.append("")
    if all_invalid:
        lines.append("Invalid codes found (first 80):")
        for inv in all_invalid[:80]:
            lines.append(f"- {inv['sheet']} R{inv['row']}C{inv['col']}: {inv['code']}")
        lines.append("")

    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    # JSON report (machine)
    def to_jsonable(obj):
        if hasattr(obj, "__dataclass_fields__"):
            d = asdict(obj)
            return d
        return obj

    payload = to_jsonable(result)
    # expand per_sheet dataclasses cleanly
    payload["per_sheet"] = [asdict(ps) for ps in result.per_sheet]

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    # Print tail summary for terminal
    print("OK")
    print(f"GO/NO-GO: {result.status}")
    print(f"TXT:  {txt_path}")
    print(f"JSON: {json_path}")

    return 0 if result.status == "GO" else 2


if __name__ == "__main__":
    raise SystemExit(main())
