#!/usr/bin/env python3
from openpyxl import load_workbook
from pathlib import Path
import re

def norm(x):
    return re.sub(r"\s+", " ", (str(x) if x is not None else "")).strip()

def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--sheet", default="ESTIMATE (INPUT)")
    ap.add_argument("--rows", type=int, default=60)
    ap.add_argument("--max-cols", type=int, default=40)
    ap.add_argument("--contains", default="amount")  # substring filter
    args = ap.parse_args()

    wb = load_workbook(Path(args.template), keep_vba=True, data_only=False)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"FATAL: sheet not found: {args.sheet}")
    ws = wb[args.sheet]

    needle = args.contains.lower()

    print("SHEET:", args.sheet)
    print("Scanning first", args.rows, "rows; max cols", args.max_cols)
    print("\n--- CANDIDATE HEADER CELLS (contains: '%s') ---" % needle)

    hits = 0
    for r in range(1, args.rows + 1):
        for c in range(1, args.max_cols + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = norm(v)
            if not s:
                continue
            if needle in s.lower():
                hits += 1
                col_letter = ws.cell(r, c).coordinate[:-len(str(r))]
                print(f"- R{r}C{c} ({ws.cell(r,c).coordinate}): {s}")

    if hits == 0:
        print("(none)")

    print("\n--- ROWS WITH 'CODE' (to locate header row) ---")
    for r in range(1, args.rows + 1):
        row_txt = " | ".join(
            [norm(ws.cell(r, c).value) for c in range(1, args.max_cols + 1) if norm(ws.cell(r,c).value)]
        )
        if "code" in row_txt.lower():
            print(f"ROW {r}: {row_txt}")

    print("\nTIP: Pick the exact header text used for the amount column and rerun patch with --amount-header \"<THAT TEXT>\"")

if __name__ == "__main__":
    main()
